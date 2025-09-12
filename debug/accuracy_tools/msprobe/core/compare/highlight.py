# Copyright (c) 2024-2025, Huawei Technologies Co., Ltd.
# All rights reserved.
#
# Licensed under the Apache License, Version 2.0  (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import abc
import math
import multiprocessing
from collections import namedtuple

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from tqdm import tqdm

from msprobe.core.common.const import CompareConst, Const
from msprobe.core.common.file_utils import save_workbook
from msprobe.core.common.log import logger
from msprobe.core.common.utils import get_header_index, CompareException
from msprobe.core.compare.utils import table_value_is_valid, gen_api_batches
from msprobe.core.compare.config import ModeConfig


def add_highlight_row_info(color_list, num, highlight_err_msg):
    for i, (existing_num, existing_err_msg) in enumerate(color_list):
        if num == existing_num:
            color_list[i][1].append(highlight_err_msg)
            return
    color_list.append((num, [highlight_err_msg]))


class HighlightCheck(abc.ABC):
    @abc.abstractmethod
    def apply(self, info, color_columns, dump_mode):
        raise NotImplementedError


class CheckOrderMagnitude(HighlightCheck):
    """检查Max diff的数量级差异"""

    def apply(self, info, color_columns, dump_mode):
        api_in, api_out, num = info
        max_diff_index = get_header_index(CompareConst.MAX_DIFF if dump_mode == Const.SUMMARY
                                          else CompareConst.MAX_ABS_ERR, dump_mode)
        max_diff_in = abs(api_in[max_diff_index])
        max_diff_out = abs(api_out[max_diff_index])
        if max_diff_in > max_diff_out or (max_diff_in <= 1 or max_diff_out <= 1):
            return
        in_order = 0 if max_diff_in < 1 else math.log10(max_diff_in)
        out_order = 0 if max_diff_out < 1 else math.log10(max_diff_out)
        if out_order - in_order >= CompareConst.ORDER_MAGNITUDE_DIFF_YELLOW:
            add_highlight_row_info(color_columns.yellow, num,
                                   "maximum absolute error of both input/parameters and output exceed 1, "
                                   "with the output larger by an order of magnitude")


class CheckOneThousandErrorRatio(HighlightCheck):
    """检查千分误差比率"""

    def apply(self, info, color_columns, dump_mode):
        api_in, api_out, num = info
        one_thousand_index = get_header_index(CompareConst.ONE_THOUSANDTH_ERR_RATIO, dump_mode)
        if (not isinstance(api_in[one_thousand_index], (float, int)) or
                not isinstance(api_out[one_thousand_index], (float, int))):
            return
        if (api_in[one_thousand_index] > CompareConst.ONE_THOUSAND_ERROR_IN_RED and
                api_out[one_thousand_index] < CompareConst.ONE_THOUSAND_ERROR_OUT_RED):
            add_highlight_row_info(color_columns.red, num,
                                   "The input/parameter's one thousandth err ratio exceeds 0.9, "
                                   "while the output's is below 0.6")
        elif api_in[one_thousand_index] - api_out[one_thousand_index] > CompareConst.ONE_THOUSAND_ERROR_DIFF_YELLOW:
            add_highlight_row_info(color_columns.yellow, num,
                                   "The output's one thousandth err ratio decreases by more than 0.1 "
                                   "compared to the input/parameter's")


class CheckCosineSimilarity(HighlightCheck):
    """检查余弦相似度"""

    def apply(self, info, color_columns, dump_mode):
        api_in, api_out, num = info
        cosine_index = get_header_index(CompareConst.COSINE, dump_mode)
        if not isinstance(api_in[cosine_index], (float, int)) or not isinstance(api_out[cosine_index], (float, int)):
            return
        if api_in[cosine_index] - api_out[cosine_index] > CompareConst.COSINE_DIFF_YELLOW:
            add_highlight_row_info(color_columns.yellow, num,
                                   "The output's cosine decreases by more than 0.1 "
                                   "compared to the input/parameter's")


class CheckMaxRelativeDiff(HighlightCheck):
    """检查最大相对差异"""

    def apply(self, info, color_columns, dump_mode):
        def get_number(data):
            """统计量相对值如果为正常百分数据，str格式并以%结尾"""
            if isinstance(data, str) and data.endswith("%"):
                return float(data[:-1]) / 100
            return data

        api_in, api_out, num = info
        max_rel_diff = get_header_index(CompareConst.MAX_RELATIVE_ERR, dump_mode)
        input_max_relative_diff = api_in[max_rel_diff]  # 内部数据，长度总是和表头一致，不会越界
        output_max_relative_diff = api_out[max_rel_diff]
        input_max_relative_diff = get_number(input_max_relative_diff)
        output_max_relative_diff = get_number(output_max_relative_diff)

        if not isinstance(output_max_relative_diff, (float, int)):
            return
        if output_max_relative_diff > CompareConst.MAX_RELATIVE_OUT_RED:
            add_highlight_row_info(color_columns.red, num, "maximum relative error exceeds 0.5")

        if not isinstance(input_max_relative_diff, (float, int)):
            return
        if (output_max_relative_diff > CompareConst.MAX_RELATIVE_OUT_YELLOW and
                input_max_relative_diff < CompareConst.MAX_RELATIVE_IN_YELLOW):
            add_highlight_row_info(color_columns.yellow, num,
                                   "The output's maximum relative error exceeds 0.1, "
                                   "while the input/parameter's is below 0.01")


class CheckOverflow(HighlightCheck):
    """检查是否存在溢出"""

    def apply(self, info, color_columns, dump_mode):
        line, num = info
        npu_max_index = get_header_index(CompareConst.NPU_MAX, dump_mode)
        npu_min_index = get_header_index(CompareConst.NPU_MIN, dump_mode)
        max_diff_index = get_header_index(CompareConst.MAX_DIFF if dump_mode == Const.SUMMARY
                                          else CompareConst.MAX_ABS_ERR, dump_mode)
        if str(line[npu_max_index]) in CompareConst.OVERFLOW_LIST or str(
                line[npu_min_index]) in CompareConst.OVERFLOW_LIST:
            add_highlight_row_info(color_columns.red, num, "maximum or minimum is nan, -inf, or inf")
            return
        # check if Max_Diff > 1e+10
        if isinstance(line[max_diff_index], (float, int)) and abs(line[max_diff_index]) > CompareConst.MAX_DIFF_RED:
            add_highlight_row_info(color_columns.red, num, "maximum absolute error exceeds 1e+10")


class CheckReqGradConsist(HighlightCheck):
    """检查requires_grad是否一致"""

    def apply(self, info, color_columns, dump_mode):
        line, num = info
        req_grad_consist_index = get_header_index(CompareConst.REQ_GRAD_CONSIST, dump_mode)
        if not line[req_grad_consist_index]:
            add_highlight_row_info(color_columns.yellow, num, "requires_grad is inconsistent")


class HighlightRules:
    """高亮规则集合，用于检查API的误差"""
    # 适用于每行的规则
    basic_rules = {
        "check_overflow": CheckOverflow()
    }
    consist_rules = {
        "check_req_grad_consist": CheckReqGradConsist()
    }

    # 用于比较输入和输出的规则
    # 真实数据检查规则
    compare_rules = {
        "check_order_magnitude": CheckOrderMagnitude(),
        "check_one_thousand_error": CheckOneThousandErrorRatio(),
        "check_cosine_similarity": CheckCosineSimilarity()
    }
    # 统计量数据检查规则
    summary_compare_rules = {
        "check_order_magnitude": CheckOrderMagnitude(),
        "check_max_relative_diff": CheckMaxRelativeDiff(),
    }


class HighLight:
    def __init__(self, mode_config: ModeConfig, rank):
        self.mode_config = mode_config
        self.rank = rank

    @staticmethod
    def check_indices_numeric(api_items, indices: list):
        """检查指定索引处的值是否都为数字类型（int 或 float）"""
        return all(isinstance(api_items[i], (float, int)) for i in indices)

    @staticmethod
    def update_highlight_err_msg(result_df, highlight_dict):
        if result_df.shape[1] <= 1:
            return

        if CompareConst.NPU_MD5 in result_df.columns:
            return

        err_msg = result_df.get(CompareConst.ERROR_MESSAGE).copy()
        red_lines_num_set = highlight_dict.get('red_rows')

        for color in ['red', 'yellow']:
            line_key = f'{color}_lines'
            lines = highlight_dict.get(line_key, [])
            for line_index, messages in lines:
                if color == 'yellow' and line_index in red_lines_num_set:
                    continue  # 如果是 yellow 行，且已被 red 行覆盖，跳过

                for msg in messages:
                    if err_msg[line_index] == '':
                        err_msg[line_index] = msg
                    else:
                        err_msg[line_index] += '\n' + msg

                if color == 'red':
                    red_lines_num_set.add(line_index)

        result_df[CompareConst.ERROR_MESSAGE] = err_msg

    @staticmethod
    def compare_result_df_convert(value):
        if not isinstance(value, (float, int)) or isinstance(value, bool):  # bool类型或者非数字类型转str
            value = f"{str(value)}\t" if str(value) in ("inf", "-inf", "nan") else str(value)
        if isinstance(value, float):
            value = f"{str(value)}\t" if str(value) in ("inf", "-inf", "nan") else value
        return value

    @staticmethod
    def value_check(value, api_name=None, i=None, result_df_columns=None):
        if not table_value_is_valid(value):
            if result_df_columns:
                logger.error(f"Malicious value [{value}] at api_name [{api_name}], column [{result_df_columns[i]}], "
                             f"is not allowed to be written into the compare result xlsx.")
            else:
                logger.error(f"Malicious value [{value}] is not allowed to be written into the compare result xlsx.")

    def find_compare_result_error_rows(self, result_df, highlight_dict):
        """将dataframe根据API分组，并找到有误差的算子用于高亮"""
        result = result_df.values
        header = result_df.columns.tolist()
        api_batches = gen_api_batches(result, header)
        default_bar_desc = 'API/Module Analyse Progress'
        bar_desc_add_rank = f'[{self.rank}]' + default_bar_desc if self.rank else default_bar_desc
        with tqdm(total=len(api_batches), desc=bar_desc_add_rank, unit="item", ncols=100) as progress_bar:
            for api_batch in api_batches:
                self.find_error_rows(result[api_batch.start: api_batch.params_grad_end_index], api_batch,
                                     highlight_dict)
                progress_bar.update(1)

    def find_error_rows(self, result, api_batch, highlight_dict):
        """找到单个API中需要高亮的行"""
        if self.mode_config.dump_mode == Const.MD5:
            return
        npu_max_index = get_header_index(CompareConst.NPU_MAX, self.mode_config.dump_mode)
        bench_max_index = get_header_index(CompareConst.BENCH_MAX, self.mode_config.dump_mode)
        max_diff_index = get_header_index(CompareConst.MAX_DIFF if self.mode_config.dump_mode == Const.SUMMARY
                                          else CompareConst.MAX_ABS_ERR, self.mode_config.dump_mode)

        red_lines, yellow_lines = [], []
        LineInfo = namedtuple('LineInfo', ['line_data', 'num_pointer'])
        ApiInfo = namedtuple('ApiInfo', ['api_input', 'api_output', 'num_pointer'])
        ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])
        color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)

        api_batch_start = api_batch.start  # result_df的input起始全局索引
        api_batch_params_end_index = api_batch.params_end_index  # result_df的params结束全局索引 + 1
        api_batch_output_end_index = api_batch.output_end_index  # result_df的output结束全局索引 + 1
        api_batch_params_slice_index_local = api_batch_params_end_index - api_batch_start  # result的params结束局部切片索引
        api_batch_output_slice_index_local = api_batch_output_end_index - api_batch_start  # result的output结束局部切片索引

        # 对单行API的输入或输出进行误差判断
        for i, line in enumerate(result):
            index = api_batch_start + i
            line_info = LineInfo(line_data=line, num_pointer=index)
            for rule in HighlightRules.basic_rules.values():
                rule.apply(line_info, color_columns, self.mode_config.dump_mode)

        # 对API的输出与输入比较，进行误差判断
        for n, api_out in enumerate(result[api_batch_params_slice_index_local: api_batch_output_slice_index_local]):
            index = api_batch_start + api_batch_params_slice_index_local + n
            # 单行检查只有溢出检查（红色），如果已经溢出，不进一步检查
            if index in red_lines:
                continue
            if not self.check_indices_numeric(api_out, [npu_max_index, bench_max_index, max_diff_index]):
                continue

            # input/parameters的比较检查, 这里api_in包括input、parameters
            for api_in in result[0: api_batch_params_slice_index_local]:
                if not self.check_indices_numeric(api_in, [npu_max_index, bench_max_index, max_diff_index]):
                    continue
                api_info = ApiInfo(api_input=api_in, api_output=api_out, num_pointer=index)
                self.apply_comparison_rules(api_info, color_columns)

        # 对单行API的输入或输出进行requires_grad是否一致判断
        for i, line in enumerate(result):
            index = api_batch_start + i
            line_info = LineInfo(line_data=line, num_pointer=index)
            for rule in HighlightRules.consist_rules.values():
                rule.apply(line_info, color_columns, self.mode_config.dump_mode)

        red_lines_num_set = {x[0] for x in red_lines}
        yellow_lines_num_set = {x[0] for x in yellow_lines}
        highlight_dict.get('red_rows', set()).update(red_lines_num_set)
        highlight_dict.get('yellow_rows', set()).update(yellow_lines_num_set - red_lines_num_set)
        highlight_dict.get('red_lines', []).extend(red_lines)
        highlight_dict.get('yellow_lines', []).extend(yellow_lines)

    def apply_comparison_rules(self, api_info, color_columns):
        """output与input/params的比较"""
        if self.mode_config.dump_mode == Const.SUMMARY:
            for rule in HighlightRules.summary_compare_rules.values():
                rule.apply(api_info, color_columns, self.mode_config.dump_mode)
        else:
            for rule in HighlightRules.compare_rules.values():
                rule.apply(api_info, color_columns, self.mode_config.dump_mode)

    def highlight_rows_xlsx(self, result_df, highlight_dict, file_path):
        """Write and highlight results in Excel"""

        self.update_highlight_err_msg(result_df, highlight_dict)  # add highlight err_msg

        self.df_malicious_value_check(result_df)

        wb = openpyxl.Workbook()
        ws = wb.active
        result_df_convert = result_df.applymap(self.compare_result_df_convert)
        for row in dataframe_to_rows(result_df_convert, index=False, header=True):
            ws.append(row)

        # 对可疑数据标色
        logger.info('Coloring Excel in progress.')
        red_fill = PatternFill(start_color=CompareConst.RED, end_color=CompareConst.RED, fill_type="solid")
        yellow_fill = PatternFill(start_color=CompareConst.YELLOW, end_color=CompareConst.YELLOW, fill_type="solid")
        col_len = len(result_df.columns)
        for i in highlight_dict.get("red_rows", []):
            for j in range(1, col_len + 1):
                ws.cell(row=i + 2, column=j).fill = red_fill  # 2因为ws.cell中的row或column需要>=1,数据从第2行开始
        for i in highlight_dict.get("yellow_rows", []):
            for j in range(1, col_len + 1):
                ws.cell(row=i + 2, column=j).fill = yellow_fill

        save_workbook(wb, file_path)

    def handle_multi_process_malicious_value_check(self, func, result_df):
        result_total_nums = len(result_df)
        process_num = int((multiprocessing.cpu_count() + 1) / 2)

        if result_total_nums <= process_num:
            process_num = 1
            chunks = [result_df]
        else:
            chunk_size = result_total_nums // process_num
            chunks = [result_df.iloc[i: i + chunk_size] for i in range(0, result_total_nums, chunk_size)]

        pool = multiprocessing.Pool(process_num)

        def err_call(args):
            logger.error("Multiprocessing malicious value check failed! Reason: {}".format(args))

        result_df_columns = result_df.columns.tolist()
        for column in result_df_columns:
            self.value_check(column)
        async_results = []
        for df_chunk in chunks:
            result = pool.apply_async(func, args=(df_chunk, result_df_columns,), error_callback=err_call)
            async_results.append(result)

        pool.close()

        for ar in async_results:
            try:
                ar.get(timeout=3600)
            except Exception as e:
                logger.error(f"Task failed with exception: {e}")
                pool.terminate()
                raise CompareException(CompareException.MULTIPROCESS_ERROR) from e

        pool.join()

    def df_malicious_value_check(self, result_df):
        result_df_columns = result_df.columns.tolist()
        for column in result_df_columns:
            self.value_check(column)
        for row in result_df.itertuples(index=False):
            api_name = row[0]
            for i, value in enumerate(row):
                self.value_check(value, api_name, i, result_df_columns)
