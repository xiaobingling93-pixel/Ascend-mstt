#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
# Copyright (C) 2019-2024. Huawei Technologies Co., Ltd. All rights reserved.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""

import json
import multiprocessing
import os.path
import sys
import torch
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from collections import namedtuple
from dataclasses import dataclass

from msprobe.pytorch.compare.match import graph_mapping
from msprobe.pytorch.compare.highlight import HighlightRules, get_header_index
from msprobe.pytorch.compare.npy_compare import compare_ops_apply, get_error_type, reshape_value, get_relative_err, \
    get_error_message
from msprobe.pytorch.advisor.advisor import Advisor
from msprobe.pytorch.common.log import logger
from msprobe.core.common.utils import check_compare_param, add_time_with_xlsx, CompareException, \
    format_value, check_file_not_exists, check_configuration_param, task_dumppath_get
from msprobe.core.common.file_check import FileChecker, change_mode, FileOpen, create_directory
from msprobe.core.common.const import Const, CompareConst, FileCheckConst
from msprobe.core.common.exceptions import FileCheckException


def check_graph_mode(a_op_name, b_op_name):
    if "Aten" in a_op_name and "Aten" not in b_op_name:
        return True
    if "Aten" not in a_op_name and "Aten" in b_op_name:
        return True
    return False


def check_op(npu_dict, bench_dict, fuzzy_match):
    a_op_name = npu_dict["op_name"]
    b_op_name = bench_dict["op_name"]
    graph_mode = check_graph_mode(a_op_name[0], b_op_name[0])
    if graph_mode:
        return graph_mapping.match(a_op_name[0], b_op_name[0])
    struct_match = check_struct_match(npu_dict, bench_dict)
    if not fuzzy_match:
        return a_op_name == b_op_name and struct_match
    is_match = True
    try:
        is_match = fuzzy_check_op(a_op_name, b_op_name)
    except Exception as err:
        logger.warning("%s and %s can not fuzzy match." % (a_op_name, b_op_name))
        is_match = False
    return is_match and struct_match


def check_struct_match(npu_dict, bench_dict):
    npu_struct_in = npu_dict.get("input_struct")
    bench_struct_in = bench_dict.get("input_struct")
    npu_struct_out = npu_dict.get("output_struct")
    bench_struct_out = bench_dict.get("output_struct")
    is_match = npu_struct_in == bench_struct_in and npu_struct_out == bench_struct_out
    if not is_match:
        if len(npu_struct_in) == 0 or len(bench_struct_in) == 0 or len(npu_struct_in) != len(bench_struct_in):
            return False
        struct_in_is_match = check_type_shape_match(npu_struct_in, bench_struct_in)
        struct_out_is_match = check_type_shape_match(npu_struct_out, bench_struct_out)
        is_match = struct_in_is_match and struct_out_is_match
    return is_match


def check_type_shape_match(npu_struct, bench_struct):
    shape_type_match = False
    for npu_type_shape, bench_type_shape in zip(npu_struct, bench_struct):
        npu_type = npu_type_shape[0]
        npu_shape = npu_type_shape[1]
        bench_type = bench_type_shape[0]
        bench_shape = bench_type_shape[1]
        shape_match = npu_shape == bench_shape
        type_match = npu_type == bench_type
        if not type_match:
            if [npu_type, bench_type] in [["torch.float16", "torch.float32"], ["torch.float32", "torch.float16"],
                                          ["torch.float16", "torch.bfloat16"], ["torch.bfloat16", "torch.float16"]]:
                type_match = True
            else:
                type_match = False
        shape_type_match = shape_match and type_match
        if not shape_type_match:
            return False
    return shape_type_match


def fuzzy_check_op(npu_name_list, bench_name_list):
    if len(npu_name_list) == 0 or len(bench_name_list) == 0 or len(npu_name_list) != len(bench_name_list):
        return False
    is_match = True
    for npu_name, bench_name in zip(npu_name_list, bench_name_list):
        is_match = fuzzy_check_name(npu_name, bench_name)
        if not is_match:
            break
    return is_match


def fuzzy_check_name(npu_name, bench_name):
    if "forward" in npu_name and "forward" in bench_name:
        is_match = rename_api(npu_name, "forward") == rename_api(bench_name, "forward")
    elif "backward" in npu_name and "backward" in bench_name:
        is_match = rename_api(npu_name, "backward") == rename_api(bench_name, "backward")
    else:
        is_match = npu_name == bench_name
    return is_match


def rename_api(npu_name, process):
    npu_split = npu_name.split(process)
    torch_func_index, in_out = npu_split[0], npu_split[1]
    torch_func_split = torch_func_index.rsplit(Const.SEP, 2)
    torch_func = str(torch_func_split[0]) + str(in_out)
    return torch_func


def merge_tensor(tensor_list, summary_compare, md5_compare):
    op_dict = {}
    op_dict["op_name"] = []
    op_dict["input_struct"] = []
    op_dict["kwargs_struct"] = []
    op_dict["output_struct"] = []
    op_dict["summary"] = []
    op_dict["stack_info"] = []

    all_mode_bool = not (summary_compare or md5_compare)
    if all_mode_bool:
        op_dict["data_name"] = []

    for tensor in tensor_list:
        if len(tensor) == 2:
            op_dict['stack_info'].append(tensor['full_info'])
            break
        op_dict["op_name"].append(tensor['full_op_name'])
        if not md5_compare:
            if tensor['full_op_name'].find("input") != -1:
                op_dict["input_struct"].append((tensor['dtype'], tensor['shape']))
            elif tensor['full_op_name'].find("kwarg") != -1:
                op_dict["kwargs_struct"].append((tensor['dtype'], tensor['shape']))
            elif tensor['full_op_name'].find("output") != -1:
                op_dict["output_struct"].append((tensor['dtype'], tensor['shape']))
        else:
            if tensor['full_op_name'].find("input") != -1:
                op_dict["input_struct"].append((tensor['dtype'], tensor['shape'], tensor['md5']))
            elif tensor['full_op_name'].find("kwarg") != -1:
                op_dict["kwargs_struct"].append((tensor['dtype'], tensor['shape'], tensor['md5']))
            elif tensor['full_op_name'].find("output") != -1:
                op_dict["output_struct"].append((tensor['dtype'], tensor['shape'], tensor['md5']))

        op_dict["summary"].append([tensor['Max'], tensor['Min'], tensor['Mean'], tensor['Norm']])

        if all_mode_bool:
            op_dict["data_name"].append(tensor['data_name'])

    if not op_dict["kwargs_struct"]:
        del op_dict["kwargs_struct"]
    return op_dict if op_dict["op_name"] else {}


def match_op(npu_queue, bench_queue, fuzzy_match):
    for b_index, b_op in enumerate(bench_queue[0: -1]):
        if check_op(npu_queue[-1], b_op, fuzzy_match):
            return len(npu_queue) - 1, b_index
    if check_op(npu_queue[-1], bench_queue[-1], fuzzy_match):
        return len(npu_queue) - 1, len(bench_queue) - 1
    for n_index, n_op in enumerate(npu_queue[0: -1]):
        if check_op(n_op, bench_queue[-1], fuzzy_match):
            return n_index, len(bench_queue) - 1
    return -1, -1


def get_accuracy(result, n_dict, b_dict, summary_compare=False, md5_compare=False):
    def get_accuracy_core(n_start, n_len, b_start, b_len, key):
        min_len = min(n_len, b_len)
        npu_stack_info = n_dict.get("stack_info", None)
        bench_stack_info = b_dict.get("stack_info", None)
        has_stack = npu_stack_info and bench_stack_info

        all_mode_bool = not (summary_compare or md5_compare)
        if all_mode_bool:
            npu_data_name = n_dict.get("data_name", None)
            bench_data_name = b_dict.get("data_name", None)

        for index in range(min_len):

            n_name = n_dict['op_name'][n_start + index]
            b_name = b_dict['op_name'][b_start + index]
            n_struct = n_dict[key][index]
            b_struct = b_dict[key][index]
            err_msg = ""
            if md5_compare:
                result_item = [n_name, b_name, n_struct[0], b_struct[0], n_struct[1], b_struct[1],
                               n_struct[2], b_struct[2],
                               CompareConst.PASS if n_struct[2] == b_struct[2] else CompareConst.DIFF]
                if has_stack and index == 0 and key == "input_struct":
                    result_item.extend(npu_stack_info)
                else:
                    result_item.append(CompareConst.NONE)
                result.append(result_item)
                continue

            if summary_compare:
                result_item = [n_name, b_name, n_struct[0], b_struct[0], n_struct[1], b_struct[1],
                               " ", " ", " ", " ", " ", " ", " ", " "]
            else:
                result_item = [n_name, b_name, n_struct[0], b_struct[0], n_struct[1], b_struct[1],
                               " ", " ", " ", " ", " "]

            npu_summary_data = n_dict.get("summary")[n_start + index]
            result_item.extend(npu_summary_data)
            bench_summary_data = b_dict.get("summary")[b_start + index]
            result_item.extend(bench_summary_data)

            if summary_compare:
                start_idx = CompareConst.SUMMARY_COMPARE_RESULT_HEADER.index(CompareConst.MAX_DIFF)
                warning_flag = False
                for i, (npu_val, bench_val) in enumerate(zip(npu_summary_data, bench_summary_data)):
                    if isinstance(npu_val, (float, int)) and isinstance(bench_val, (float, int)):
                        diff = npu_val - bench_val
                        if bench_val != 0:
                            relative = str(abs((diff / bench_val) * 100)) + '%'
                        else:
                            relative = "N/A"
                        result_item[start_idx + i] = diff
                        result_item[start_idx + i + 4] = relative
                        magnitude_diff = abs(diff) / (max(abs(npu_val), abs(bench_val)) + 1e-10)
                        if magnitude_diff > 0.5:
                            warning_flag = True
                    else:
                        result_item[start_idx + i] = CompareConst.NONE
                accuracy_check = CompareConst.WARNING if warning_flag else ""
                err_msg += "Need double check api accuracy." if warning_flag else ""
                for i in range(start_idx, len(result_item)):
                    if str(result_item[i]) in ('inf', '-inf', 'nan'):
                        result_item[i] = f'{result_item[i]}\t'

            result_item.append(accuracy_check if summary_compare else CompareConst.ACCURACY_CHECK_YES)
            result_item.append(err_msg)
            if has_stack and index == 0 and key == "input_struct":
                result_item.extend(npu_stack_info)
            else:
                result_item.append(CompareConst.NONE)
            if all_mode_bool:
                result_item.append(npu_data_name[n_start + index])

            result.append(result_item)

        if n_len > b_len:
            for index in range(b_len, n_len):
                n_name = n_dict['op_name'][n_start + index]
                n_struct = n_dict[key][index]
                if md5_compare:
                    result_item = [n_name, CompareConst.NAN, n_struct[0], CompareConst.NAN,
                                   n_struct[1], CompareConst.NAN, n_struct[2], CompareConst.NAN, CompareConst.NAN]
                    result.append(result_item)
                    continue
                result_item = [n_name, CompareConst.NAN, n_struct[0], CompareConst.NAN,
                               n_struct[1], CompareConst.NAN, " ", " ", " ", " ", " "]
                summary_data = n_dict.get("summary")[n_start + index]
                result_item.extend(summary_data)
                summary_data = [CompareConst.NAN for _ in range(len(n_dict.get("summary")[0]))]
                result_item.extend(summary_data)

                err_msg = ""
                result_item.append(CompareConst.ACCURACY_CHECK_YES)
                result_item.append(err_msg)

                if has_stack and index == 0 and key == "input_struct":
                    result_item.extend(npu_stack_info)
                else:
                    result_item.append(CompareConst.NONE)
                if all_mode_bool:
                    result_item.append(npu_data_name[n_start + index])

                result.append(result_item)

    n_num = len(n_dict['op_name'])
    b_num = len(b_dict['op_name'])
    n_num_input = len([name for name in n_dict['op_name'] if 'input' in name])
    b_num_input = len([name for name in b_dict['op_name'] if 'input' in name])
    n_num_kwarg = len([name for name in n_dict['op_name'] if 'kwarg' in name])
    b_num_kwarg = len([name for name in b_dict['op_name'] if 'kwarg' in name])
    n_num_output = n_num - n_num_input - n_num_kwarg
    b_num_output = b_num - b_num_input - b_num_kwarg
    get_accuracy_core(0, n_num_input, 0, b_num_input, 'input_struct')
    get_accuracy_core(n_num_input, n_num_kwarg, b_num_input, b_num_kwarg, "kwargs_struct")
    get_accuracy_core(n_num_input + n_num_kwarg, n_num_output, b_num_input + b_num_kwarg, b_num_output, 'output_struct')


def _do_multi_process(input_parma, result_df):
    try:
        result_df = _handle_multi_process(compare_ops, input_parma, result_df, multiprocessing.Manager().RLock())
        return result_df
    except ValueError as e:
        logger.error('result dataframe is not found.')
        raise CompareException(CompareException.INVALID_DATA_ERROR) from e


def read_dump_data(result_df):
    try:
        npu_dump_name_list = result_df.iloc[0:, 0].tolist()
        npu_dump_tensor_list = result_df.iloc[0:, -1].tolist()
        op_name_mapping_dict = {}
        for index, _ in enumerate(npu_dump_name_list):
            npu_dump_name = npu_dump_name_list[index]
            npu_dump_tensor = npu_dump_tensor_list[index]
            op_name_mapping_dict[npu_dump_name] = [npu_dump_tensor, npu_dump_tensor]
        return op_name_mapping_dict
    except ValueError as e:
        logger.error('result dataframe is not found.')
        raise CompareException(CompareException.INVALID_DATA_ERROR) from e
    except IndexError as e:
        logger.error('result dataframe elements can not be access.')
        raise CompareException(CompareException.INDEX_OUT_OF_BOUNDS_ERROR) from e


def _handle_multi_process(func, input_parma, result_df, lock):
    process_num = int((multiprocessing.cpu_count() + 1) / 2)
    op_name_mapping_dict = read_dump_data(result_df)

    df_chunk_size = len(result_df) // process_num
    if df_chunk_size > 0:
        df_chunks = [result_df.iloc[i:i + df_chunk_size] for i in range(0, len(result_df), df_chunk_size)]
    else:
        df_chunks = [result_df]

    results = []
    pool = multiprocessing.Pool(process_num)

    def err_call(args):
        logger.error('multiprocess compare failed! Reason: {}'.format(args))
        try:
            pool.terminate()
        except OSError as e:
            logger.error("pool terminate failed")

    for process_idx, df_chunk in enumerate(df_chunks):
        idx = df_chunk_size * process_idx
        result = pool.apply_async(func,
                                  args=(idx, op_name_mapping_dict, df_chunk, lock, input_parma),
                                  error_callback=err_call)
        results.append(result)
    final_results = [r.get() for r in results]
    pool.close()
    pool.join()
    return pd.concat(final_results, ignore_index=True)


def compare_ops(idx, dump_path_dict, result_df, lock, input_parma):
    cos_result = []
    max_err_result = []
    max_relative_err_result = []
    err_mess = []
    one_thousand_err_ratio_result = []
    five_thousand_err_ratio_result = []
    is_print_compare_log = input_parma.get("is_print_compare_log")
    for i in range(len(result_df)):
        op_name = result_df.iloc[i, 0]
        if is_print_compare_log:
            logger.info("start compare: {}".format(op_name))
        cos_sim, max_abs_err, max_relative_err, one_thousand_err_ratio, five_thousand_err_ratio, err_msg = compare_by_op(
            op_name, dump_path_dict, input_parma)
        if is_print_compare_log:
            logger.info(
                "[{}] Compare result: cosine {}, max_abs_err {}, max_relative_err {}, {}, one_thousand_err_ratio {}, "
                "five_thousand_err_ratio {}".format(op_name, cos_sim, max_abs_err, max_relative_err, err_msg,
                                                    one_thousand_err_ratio, five_thousand_err_ratio))
        cos_result.append(cos_sim)
        max_err_result.append(max_abs_err)
        max_relative_err_result.append(max_relative_err)
        err_mess.append(err_msg)
        one_thousand_err_ratio_result.append(one_thousand_err_ratio)
        five_thousand_err_ratio_result.append(five_thousand_err_ratio)

    cr = ComparisonResult(
        cos_result=cos_result,
        max_err_result=max_err_result,
        max_relative_err_result=max_relative_err_result,
        err_msgs=err_mess,
        one_thousand_err_ratio_result=one_thousand_err_ratio_result,
        five_thousand_err_ratio_result=five_thousand_err_ratio_result
    )

    return _save_cmp_result(idx, cr, result_df, lock)


@dataclass
class ComparisonResult:
    cos_result: list
    max_err_result:  list
    max_relative_err_result: list
    err_msgs: list
    one_thousand_err_ratio_result: list
    five_thousand_err_ratio_result: list


def _save_cmp_result(offset, result: ComparisonResult, result_df, lock):
    """
        Save comparison results into the result DataFrame with thread safety.
    Args:
        offset: offset for index
        result: data struct of ComparisonResult
        result_df: result of DataFrame
        lock: thread lock

    Returns:
        comparison results in DataFrame
    """

    lock.acquire()
    try:
        for i, _ in enumerate(result.cos_result):
            process_index = i + offset
            result_df.loc[process_index, CompareConst.COSINE] = result.cos_result[i]
            result_df.loc[process_index, CompareConst.MAX_ABS_ERR] = result.max_err_result[i]
            result_df.loc[process_index, CompareConst.MAX_RELATIVE_ERR] = result.max_relative_err_result[i]
            result_df.loc[process_index, CompareConst.ERROR_MESSAGE] = result.err_msgs[i]
            result_df.loc[process_index, CompareConst.ACCURACY] = check_accuracy(result.cos_result[i], result.max_err_result[i])
            result_df.loc[process_index, CompareConst.ONE_THOUSANDTH_ERR_RATIO] = result.one_thousand_err_ratio_result[i]
            result_df.loc[process_index, CompareConst.FIVE_THOUSANDTHS_ERR_RATIO] = result.five_thousand_err_ratio_result[i]
        return result_df
    except ValueError as e:
        logger.error('result dataframe is not found.')
        raise CompareException(CompareException.INVALID_DATA_ERROR) from e
    except IndexError as e:
        logger.error('result dataframe elements can not be access.')
        raise CompareException(CompareException.INDEX_OUT_OF_BOUNDS_ERROR) from e
    finally:
        lock.release()


def check_accuracy(cos, max_abs_err):
    if cos == CompareConst.SHAPE_UNMATCH:
        return CompareConst.ACCURACY_CHECK_UNMATCH
    if cos == CompareConst.NONE or max_abs_err == CompareConst.NONE:
        return CompareConst.NONE
    if cos == "N/A" or max_abs_err == "N/A":
        return CompareConst.ACCURACY_CHECK_NO
    try:
        cos, max_abs_err = float(cos), float(max_abs_err)
    except ValueError:
        logger.warning("Cosine or MaxAbsErr can not get float value.")
        return CompareConst.NONE
    if cos < CompareConst.COS_THRESHOLD and max_abs_err > CompareConst.MAX_ABS_ERR_THRESHOLD:
        return CompareConst.ACCURACY_CHECK_NO
    if cos < CompareConst.COS_MAX_THRESHOLD or max_abs_err > CompareConst.MAX_ABS_ERR_MAX_THRESHOLD:
        return CompareConst.ACCURACY_CHECK_NO
    return CompareConst.ACCURACY_CHECK_YES


def read_npy_data(dir_path, file_name):
    data_path = os.path.join(dir_path, file_name)
    path_checker = FileChecker(data_path, FileCheckConst.FILE, FileCheckConst.READ_ABLE,
                               FileCheckConst.PT_SUFFIX, False)
    data_path = path_checker.common_check()
    data_value = torch.load(data_path, map_location=torch.device('cpu')).detach()       # detach for less memory
    if data_value.dtype == torch.bfloat16:
        data_value = data_value.to(torch.float32)
    data_value = data_value.numpy()
    return data_value


def compare_by_op(op_name, op_name_mapping_dict, input_parma):
    npu_bench_name_list = op_name_mapping_dict[op_name]
    data_name = npu_bench_name_list[1]
    error_file, relative_err, error_flag = None, None, False
    if data_name == '-1' or data_name == -1:  # 没有真实数据路径
        n_value, b_value = CompareConst.READ_NONE, CompareConst.READ_NONE
        error_flag = True
    else:
        try:
            n_value = read_npy_data(input_parma.get("npu_dump_data_dir"), npu_bench_name_list[0])
            b_value = read_npy_data(input_parma.get("bench_dump_data_dir"), npu_bench_name_list[1])
        except IOError as error:
            error_file = error.filename
            n_value, b_value = CompareConst.READ_NONE, CompareConst.READ_NONE
            error_flag = True
        except FileCheckException:
            error_file = data_name
            n_value, b_value = CompareConst.READ_NONE, CompareConst.READ_NONE
            error_flag = True

    n_value, b_value, error_flag = get_error_type(n_value, b_value, error_flag)
    if not error_flag:
        relative_err = get_relative_err(n_value, b_value)
        n_value, b_value = reshape_value(n_value, b_value)

    err_msg = get_error_message(n_value, b_value, op_name, error_flag, error_file=error_file)
    result_list, err_msg = compare_ops_apply(n_value, b_value, error_flag, err_msg, relative_err=relative_err)

    if npu_bench_name_list[0] != npu_bench_name_list[1]:
        err_msg += " Fuzzy matching data, the comparison accuracy may be affected."
    result_list.append(err_msg)
    return result_list


def handle_inf_nan(n_value, b_value):
    n_inf = np.isinf(n_value)
    b_inf = np.isinf(b_value)
    n_nan = np.isnan(n_value)
    b_nan = np.isnan(b_value)

    # merge boolean expressions
    any_inf = np.any(n_inf) or np.any(b_inf)
    any_nan = np.any(n_nan) or np.any(b_nan)
    if any_inf or any_nan:
        if np.array_equal(n_inf, b_inf) and np.array_equal(n_nan, b_nan):
            n_value[n_inf] = 0
            b_value[b_inf] = 0
            n_value[n_nan] = 0
            b_value[b_nan] = 0
        else:
            return CompareConst.NAN, CompareConst.NAN
    return n_value, b_value


def find_error_rows(result, last_len, n_num_input, highlight_dict, summary_compare=False, md5_compare=False):
    """找到单个API中需要高亮的行"""
    if md5_compare:
        return
    npu_max_index = get_header_index('NPU max', summary_compare)
    bench_max_index = get_header_index('Bench max', summary_compare)
    max_diff_index = get_header_index('Max diff' if summary_compare else 'MaxAbsErr', summary_compare)

    red_lines, yellow_lines = [], []
    LineInfo = namedtuple('LineInfo', ['line_data', 'num_pointer'])
    ApiInfo = namedtuple('ApiInfo', ['api_input', 'api_output', 'num_pointer'])
    ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])
    color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)

    # 对单行API的输入或输出进行误差判断
    for i, line in enumerate(result):
        num = last_len + i
        line_info = LineInfo(line_data=line, num_pointer=num)
        for rule in HighlightRules.basic_rules.values():
            rule.apply(line_info, color_columns, summary_compare)

    # 对API的输出与输入比较，进行误差判断
    for n, api_out in enumerate(result[n_num_input:len(result)]):
        num = last_len + n_num_input + n
        if num in red_lines:
            continue
        if not isinstance(api_out[npu_max_index], (float, int)) \
                or not isinstance(api_out[bench_max_index], (float, int)) \
                or not isinstance(api_out[max_diff_index], (float, int)):
            continue
        for _, api_in in enumerate(result[0:n_num_input]):
            if not isinstance(api_in[npu_max_index], (float, int)) \
                    or not isinstance(api_in[bench_max_index], (float, int)) \
                    or not isinstance(api_in[max_diff_index], (float, int)):
                continue

            api_info = ApiInfo(api_input=api_in, api_output=api_out, num_pointer=num)
            if summary_compare:
                for rule in HighlightRules.summary_compare_rules.values():
                    rule.apply(api_info, color_columns, summary_compare)
            else:
                for rule in HighlightRules.compare_rules.values():
                    rule.apply(api_info, color_columns, summary_compare)

    highlight_dict.get('red_rows', []).extend(list(set(red_lines)))
    highlight_dict.get('yellow_rows', []).extend(list(set(yellow_lines) - set(red_lines)))


def get_name_and_state(name):
    """Get api/module name and state"""
    if "input" in name:
        api_name = name.split("input")[0]
        state = "input"
    else:
        api_name = name.split("output")[0]
        state = "output"
    return api_name, state


def find_compare_result_error_rows(result_df, highlight_dict, summary_compare, md5_compare):
    """将dataframe根据API分组，并找到有误差的算子用于高亮"""
    result = result_df.values
    start, input_num, output_num, end = 0, 0, 0, len(result_df)
    last_api_name, last_state = None, None
    num, last_len = 0, 0
    for res_i in result:
        api_name, state = get_name_and_state(res_i[0])
        if last_api_name:
            if api_name == last_api_name:
                if state == last_state:
                    num += 1
                else:
                    input_num = num
                    num, last_state = 1, state
            else:
                output_num = num
                find_error_rows(result[start:start + input_num + output_num], start, input_num, highlight_dict,
                                summary_compare, md5_compare)
                num, last_api_name, last_state = 1, api_name, state
                start += input_num + output_num
                input_num, output_num = 1, 0
        else:
            num, last_api_name, last_state = 1, api_name, state
    if state:
        if state == "input":
            input_num = num
        else:
            output_num = num
        find_error_rows(result[start:start + input_num + output_num], start, input_num, highlight_dict, summary_compare, md5_compare)


def highlight_rows_xlsx(result_df, highlight_dict, file_path):
    """Write and highlight results in Excel"""
    logger.info('Compare result is %s' % file_path)

    wb = openpyxl.Workbook()
    ws = wb.active

    # write header
    for j, col_name in enumerate(result_df.columns, start=1):
        ws.cell(row=1, column=j, value=col_name)

    for i, row in enumerate(result_df.iterrows(), start=2):
        for j, value in enumerate(row[1], start=1):
            if not isinstance(value, (float, int)):
                value = f'{str(value)}\t' if str(value) in ('inf', '-inf', 'nan') else str(value)
            ws.cell(row=i, column=j, value=f'{str(value)}\t' if str(value) in ('inf', '-inf', 'nan') else value)

            if (i - 2) in highlight_dict['red_rows']:
                ws.cell(row=i, column=j).fill = PatternFill(start_color=CompareConst.RED,
                                                            end_color=CompareConst.RED, fill_type="solid")
            elif (i - 2) in highlight_dict['yellow_rows']:
                ws.cell(row=i, column=j).fill = PatternFill(start_color=CompareConst.YELLOW,
                                                            end_color=CompareConst.YELLOW, fill_type="solid")
    try:
        wb.save(file_path)
    except Exception as e:
        logger.error('Save result file failed')
        raise CompareException(CompareException.WRITE_FILE_ERROR) from e
    change_mode(file_path, FileCheckConst.DATA_FILE_AUTHORITY)


def compare(input_parma, output_path, stack_mode=False, auto_analyze=True,
            fuzzy_match=False):
    try:
        summary_compare, md5_compare = task_dumppath_get(input_parma)
        check_configuration_param(stack_mode, auto_analyze, fuzzy_match)
        create_directory(output_path)
        check_compare_param(input_parma, output_path, stack_mode, summary_compare, md5_compare)
    except CompareException as error:
        logger.error('Compare failed. Please check the arguments and do it again!')
        sys.exit(error.code)
    compare_core(input_parma, output_path, stack_mode=stack_mode,
                 auto_analyze=auto_analyze, fuzzy_match=fuzzy_match, summary_compare=summary_compare,
                 md5_compare=md5_compare)


def compare_core(input_parma, output_path, **kwargs):
    """
    Compares data from multiple JSON files and generates a comparison report.

    Args:
        input_parma (dict): A dictionary containing paths to JSON files ("npu_json_path", "bench_json_path",
                            "stack_json_path").
        output_path (str): The path where the output Excel report will be saved.
        **kwargs: Additional keyword arguments including:
        - stack_mode (bool, optional): Enables stack mode comparison. Defaults to False.
        - auto_analyze (bool, optional): If True, triggers automatic analysis after comparison. Defaults to True.
        - suffix (str, optional): Suffix to append to the output file name. Defaults to ''.
        - fuzzy_match (bool, optional): Enables fuzzy matching during comparison. Defaults to False.
        - summary_compare (bool, optional): Enables summary comparison mode. Defaults to False.
        - md5_compare (bool, optional): Enables MD5 comparison. Defaults to False.

    Returns:
    """
    # get kwargs or set default value
    stack_mode = kwargs.get('stack_mode', False)
    auto_analyze = kwargs.get('auto_analyze', True)
    suffix = kwargs.get('suffix', '')
    fuzzy_match = kwargs.get('fuzzy_match', False)
    summary_compare = kwargs.get('summary_compare', False)
    md5_compare = kwargs.get('md5_compare', False)

    logger.info("Please check whether the input data belongs to you. If not, there may be security risks.")
    file_name = add_time_with_xlsx("compare_result" + suffix)
    file_path = os.path.join(os.path.realpath(output_path), file_name)
    check_file_not_exists(file_path)
    highlight_dict = {'red_rows': [], 'yellow_rows': []}

    with FileOpen(input_parma.get("npu_json_path"), "r") as npu_json, \
            FileOpen(input_parma.get("bench_json_path"), "r") as bench_json, \
            FileOpen(input_parma.get("stack_json_path"), "r") as stack_json:
        result_df = compare_process([npu_json, bench_json, stack_json], stack_mode, fuzzy_match,
                                    summary_compare, md5_compare)

    if not md5_compare and not summary_compare:
        result_df = _do_multi_process(input_parma, result_df)
    find_compare_result_error_rows(result_df, highlight_dict, summary_compare, md5_compare)
    highlight_rows_xlsx(result_df, highlight_dict, file_path)
    if auto_analyze:
        advisor = Advisor(result_df, output_path)
        advisor.analysis()


def parse(pkl_file, module_name_prefix):
    if not isinstance(module_name_prefix, str):
        logger.error("The parameter:module_name_prefix is not a string.")
        raise CompareException(CompareException.INVALID_PARAM_ERROR)
    with FileOpen(pkl_file, "r") as f:
        done = False
        title_printed = False
        while not done:
            pkl_line = f.readline()
            if pkl_line == '\n':
                continue
            if len(pkl_line) == 0:
                done = True
                break

            msg = json.loads(pkl_line)
            info_prefix = msg[0]
            if not info_prefix.startswith(module_name_prefix):
                continue

            if info_prefix.find("stack_info") != -1:
                logger.info("\nTrace back({}):".format(msg[0]))
                for item in reversed(msg[1]):
                    logger.info("  File \"{}\", line {}, in {}".format(item[0], item[1], item[2]))
                    logger.info("    {}".format(item[3]))
                continue
            if len(msg) > 5:
                summary_info = "  [{}][dtype: {}][shape: {}][max: {}][min: {}][mean: {}]" \
                    .format(msg[0], msg[3], msg[4], msg[5][0], msg[5][1], msg[5][2])
                if not title_printed:
                    logger.info("\nStatistic Info:")
                    title_printed = True
                logger.info(summary_info)


def op_item_parse(item, op_name, index, item_list=None, top_bool=True):
    if item_list is None:
        item_list = []
    if item is None or (isinstance(item, dict) and not item):
        if not top_bool:
            tmp = {'full_op_name': op_name + '.' + str(index), 'Max': None, 'Min': None, 'Mean': None, 'Norm': None,
                   'dtype': None, 'shape': None, 'md5': None, 'data_name': '-1'}
        else:
            tmp = {'full_op_name': op_name + '.0', 'Max': None, 'Min': None, 'Mean': None, 'Norm': None, 'dtype': None,
                   'shape': None, 'md5': None, 'data_name': '-1'}
        item_list.append(tmp)
        return item_list
    if index is None:
        if isinstance(item, dict):
            full_op_name = op_name + '.0'
        else:
            full_op_name = op_name
    else:
        full_op_name = op_name + Const.SEP + str(index)
    if isinstance(item, dict):
        if 'type' not in item:
            for kwarg in item:
                kwarg_parsed_list = op_item_parse(item[kwarg], op_name + Const.SEP + kwarg, None)
                item_list += kwarg_parsed_list
                kwarg_parsed_list.clear()
        elif 'dtype' in item:
            parsed_item = item
            parsed_item['full_op_name'] = full_op_name
            item_list.append(parsed_item)
        elif 'type' in item:
            parsed_item = {}
            if item['type'] == 'torch.Size':
                parsed_item['full_op_name'] = full_op_name
                parsed_item['dtype'] = 'torch.Size'
                parsed_item['shape'] = str(item['value'])
                parsed_item['md5'] = None
                parsed_item['Max'] = None
                parsed_item['Min'] = None
                parsed_item['Mean'] = None
                parsed_item['Norm'] = None
                parsed_item['data_name'] = '-1'
                item_list.append(parsed_item)
            elif item['type'] == 'slice':
                parsed_item['full_op_name'] = full_op_name
                parsed_item['dtype'] = 'slice'
                parsed_item['shape'] = str(np.shape(np.array(item['value'])))
                parsed_item['md5'] = None
                parsed_item['Max'] = None
                parsed_item['Min'] = None
                parsed_item['Mean'] = None
                parsed_item['Norm'] = None
                parsed_item['data_name'] = '-1'
                item_list.append(parsed_item)
            else:
                parsed_item['full_op_name'] = full_op_name
                parsed_item['dtype'] = str(type(item['value']))
                parsed_item['shape'] = '[]'
                parsed_item['md5'] = None
                parsed_item['Max'] = item['value']
                parsed_item['Min'] = item['value']
                parsed_item['Mean'] = item['value']
                parsed_item['Norm'] = item['value']
                parsed_item['data_name'] = '-1'
                item_list.append(parsed_item)
        else:
            resolve_api_special_parameters(item, full_op_name, item_list)
    else:
        for j, item_spec in enumerate(item):
            op_item_parse(item_spec, full_op_name, j, item_list=item_list, top_bool=False)
    return item_list


def resolve_api_special_parameters(data_dict, full_op_name, item_list):
    """
    Function Description:
        解析下面格式的数据, 是api参数的一种特殊格式
        {
         "last_hidden_state": {
          "type": "torch.Tensor",
          "dtype": "torch.bfloat16",
          ...
         },
         "loss": {
          "type": "torch.Tensor",
          "dtype": "torch.float32",
          ...
         }
        }
    Parameter:
        data_dict: 字典格式的数据
        full_op_name: 参数的全名字符串
        item_list: 参数信息集合        
    """
    for key, value in data_dict.items():
        if isinstance(value, dict):
            parsed_item = value
            parts = full_op_name.split(".")
            parts.insert(-1, key)
            full_op_name_new = ".".join(parts)
            parsed_item['full_op_name'] = full_op_name_new
            item_list.append(parsed_item)


def read_op(op_data, op_name):
    op_parsed_list = []
    if 'forward' in op_name:
        if 'input_args' in op_data:
            input_item = op_data['input_args']
            input_parsed_list = op_item_parse(input_item, op_name + '_input', None)
            op_parsed_list = input_parsed_list.copy()
            input_parsed_list.clear()
        if 'input_kwargs' in op_data:
            kwargs_item = op_data['input_kwargs']
            if isinstance(kwargs_item, dict) and "type" in kwargs_item or isinstance(kwargs_item, list):
                kwarg_parsed_list = op_item_parse(kwargs_item, op_name + '_input', None)
                op_parsed_list += kwarg_parsed_list
                kwarg_parsed_list.clear()
            elif kwargs_item:
                for kwarg in kwargs_item:
                    kwarg_parsed_list = op_item_parse(kwargs_item[kwarg], op_name + '_input.' + kwarg, None)
                    op_parsed_list += kwarg_parsed_list
                    kwarg_parsed_list.clear()
        if 'output' in op_data:
            output_item = op_data['output']
            output_parsed_list = op_item_parse(output_item, op_name + '_output', None)
            op_parsed_list += output_parsed_list
            output_parsed_list.clear()
    if 'backward' in op_name:
        if 'grad_input' in op_data:
            input_item = op_data['grad_input']
            input_parsed_list = op_item_parse(input_item, op_name + '_input', None)
            op_parsed_list = input_parsed_list.copy()
            input_parsed_list.clear()
        if 'grad_output' in op_data:
            output_item = op_data['grad_output']
            output_parsed_list = op_item_parse(output_item, op_name + '_output', None)
            op_parsed_list += output_parsed_list
            output_parsed_list.clear()
    return op_parsed_list


def compare_process(file_handles, stack_mode, fuzzy_match, summary_compare=False, md5_compare=False):
    npu_json_handle, bench_json_handle, stack_json_handle = file_handles
    npu_json_data = json.load(npu_json_handle)
    bench_json_data = json.load(bench_json_handle)
    stack_json_data = json.load(stack_json_handle)

    if fuzzy_match:
        logger.warning("This task uses fuzzy matching, which may affect the accuracy of the comparison.")

    npu_ops_queue = []
    bench_ops_queue = []
    result = []

    ops_npu_iter = iter(npu_json_data['data'])
    ops_bench_iter = iter(bench_json_data['data'])
    read_err_npu = True
    read_err_bench = True
    last_npu_ops_len = 0
    last_bench_ops_len = 0

    while True:
        if not read_err_npu and not read_err_bench:
            break
        try:
            last_npu_ops_len = len(npu_ops_queue)
            op_name_npu = next(ops_npu_iter)
            read_err_npu = True

            npu_op_data = npu_json_data['data'][op_name_npu]
            npu_op_parsed_list = read_op(npu_op_data, op_name_npu)
            if op_name_npu in stack_json_data:
                npu_op_parsed_list.append({'full_op_name': op_name_npu, 'full_info': stack_json_data[op_name_npu]})
            else:
                npu_op_parsed_list.append({'full_op_name': op_name_npu, 'full_info': None})

            npu_merge_list = merge_tensor(npu_op_parsed_list, summary_compare, md5_compare)
            if npu_merge_list:
                npu_ops_queue.append(npu_merge_list)
        except StopIteration:
            read_err_npu = False
        try:
            last_bench_ops_len = len(bench_ops_queue)
            op_name_bench = next(ops_bench_iter)

            bench_op_data = bench_json_data['data'][op_name_bench]
            bench_op_parsed_list = read_op(bench_op_data, op_name_bench)
            if op_name_bench in stack_json_data:
                bench_op_parsed_list.append(
                    {'full_op_name': op_name_bench, 'full_info': stack_json_data[op_name_bench]})
            else:
                bench_op_parsed_list.append({'full_op_name': op_name_bench, 'full_info': None})

            bench_merge_list = merge_tensor(bench_op_parsed_list, summary_compare, md5_compare)
            if bench_merge_list:
                bench_ops_queue.append(bench_merge_list)
        except StopIteration:
            read_err_bench = False

        # merge all boolean expressions
        both_empty = not npu_ops_queue and not bench_ops_queue
        no_change = (len(npu_ops_queue) == last_npu_ops_len) and (len(bench_ops_queue) == last_bench_ops_len)
        if both_empty or no_change:
            continue

        n_match_point, b_match_point = match_op(npu_ops_queue, bench_ops_queue, fuzzy_match)
        if n_match_point == -1 and b_match_point == -1:
            continue
        n_match_data = npu_ops_queue[n_match_point]
        b_match_data = bench_ops_queue[b_match_point]
        un_match_data = npu_ops_queue[0: n_match_point]
        for npu_data in un_match_data:
            get_un_match_accuracy(result, npu_data, md5_compare, summary_compare)
        get_accuracy(result, n_match_data, b_match_data, summary_compare, md5_compare)
        del npu_ops_queue[0: n_match_point + 1]
        del bench_ops_queue[0: b_match_point + 1]
    if npu_ops_queue:
        for npu_data in npu_ops_queue:
            get_un_match_accuracy(result, npu_data, md5_compare, summary_compare)

    header = []
    if md5_compare:
        header = CompareConst.MD5_COMPARE_RESULT_HEADER[:]
    elif summary_compare:
        header = CompareConst.SUMMARY_COMPARE_RESULT_HEADER[:]
    else:
        header = CompareConst.COMPARE_RESULT_HEADER[:]

    all_mode_bool = not (summary_compare or md5_compare)
    if stack_mode:
        if all_mode_bool:
            header.append(CompareConst.STACK)
            header.append(CompareConst.DATA_NAME)
        else:
            header.append(CompareConst.STACK)
    else:
        if all_mode_bool:
            for row in result:
                del row[-2]
            header.append(CompareConst.DATA_NAME)
        else:
            for row in result:
                del row[-1]

    result_df = pd.DataFrame(result, columns=header)
    return result_df


def get_un_match_accuracy(result, n_dict, md5_compare, summary_compare):
    index_out = 0
    npu_stack_info = n_dict.get("stack_info", None)
    bench_name, bench_type, bench_shape = CompareConst.NAN, CompareConst.NAN, CompareConst.NAN
    err_msg = CompareConst.NO_BENCH
    accuracy_check_res = CompareConst.NAN
    for index, n_name in enumerate(n_dict["op_name"]):
        if n_name.find("input") != -1:
            n_struct = n_dict["input_struct"][index]
        else:
            n_struct = n_dict["output_struct"][index_out]
            index_out += 1

        result_item = [n_name, bench_name, n_struct[0], bench_type, n_struct[1], bench_shape]
        if md5_compare:
            result_item.extend([CompareConst.NAN] * 3)
            if npu_stack_info and index == 0:
                result_item.extend(npu_stack_info)
            result.append(result_item)
            continue
        if summary_compare:
            result_item.extend([CompareConst.NAN] * 8)
        else:
            result_item.extend([CompareConst.NAN] * 5)
        summary_data = n_dict.get("summary")[index]
        result_item.extend(summary_data)
        summary_data = [CompareConst.NAN] * 4
        result_item.extend(summary_data)
        result_item.append(accuracy_check_res)
        result_item.append(err_msg)
        if npu_stack_info and index == 0:
            result_item.extend(npu_stack_info)
        if not md5_compare and not summary_compare and result_item[1] == CompareConst.NAN:
            if index == 0:
                result_item.extend(["-1"])
            else:
                result_item.extend([CompareConst.NONE, "-1"])
        result.append(result_item)
