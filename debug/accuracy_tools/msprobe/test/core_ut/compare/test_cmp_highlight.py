# coding=utf-8
import os
import shutil
import sys
from collections import namedtuple
import unittest
from unittest.mock import patch

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from msprobe.core.common.const import CompareConst, Const
from msprobe.core.compare.highlight import CheckMaxRelativeDiff, CheckOrderMagnitude, \
    CheckOneThousandErrorRatio, CheckCosineSimilarity, add_highlight_row_info, HighLight
from msprobe.core.compare.config import ModeConfig
from msprobe.core.compare.utils import ApiBatch


summary_line_input = ['Functional_batch_norm_0_forward.input.0', 'Functional_batch_norm_0_forward.input.0',
                      'torch.float16', 'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
                      0.01, 0, 0, 0, '0.0%', '0.0%', '0.0%', '0.0%', 1, 1, 1, 1, 1.01, 1, 1, 1,
                      True, 'Yes', '']
summary_line_1 = ['Functional_batch_norm_0_forward.output.0', 'Functional_batch_norm_0_forward.output.0',
                  'torch.float16', 'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
                  10, 0, 0, 0, '0.0%', '0.0%', '0.0%', '0.0%', 2, 0, 1, 1, 1, 1, 1, 1,
                  True, 'Warning', '']
summary_line_2 = ['Functional_batch_norm_0_forward.output.1', 'Functional_batch_norm_0_forward.output.1',
                  'torch.float16', 'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
                  0.02, 0, 0, 0, '0.0%', '0.0%', '0.0%', '0.0%', 0.12, 0, 1, 1, 0.1, 1, 1, 1,
                  True, 'Warning', '']
summary_line_3 = ['Functional_batch_norm_0_forward.output.2', 'Functional_batch_norm_0_forward.output.2',
                  'torch.float16', 'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
                  0, 0, 0, 0, '0.0%', '0.0%', '0.0%', '0.0%', 2, 0, 1, 1, 1, 1, 1, 1,
                  True, 'Warning', '']
line_input = ['Functional.batch.norm.0.forward.input.0', 'Functional.batch.norm.0.forward.input.0', 'torch.float16',
              'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
              1, 0.5, 1, 1, 0.95, 1, 1, 1, 1, 1, 1.01, 1, 1, 1,
              True, 'Yes', '', 'input', 'Functional.batch.norm.0.forward']
line_1 = ['Functional.batch.norm.0.forward.output.0', 'Functional.batch.norm.0.forward.output.0', 'torch.float16',
          'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
          0.8, 0.5, 1, 1, 0.59, 1, 'nan', 0, 1, 1, 19, 1, 1, 1,
          True, 'Yes', '', 'output', 'Functional.batch.norm.0.forward']
line_2 = ['Functional.batch.norm.0.forward.output.1', 'Functional.batch.norm.0.forward.output.1', 'torch.float16',
          'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
          0.9, 0.5, 1, 1, 0.8, 1, 0, 0.12, 0, 1, 1, 0.1, 1, 1,
          True, 'Yes', '', 'output', 'Functional.batch.norm.0.forward']
line_3 = ['Functional.batch.norm.0.forward.output.2', 'Functional.batch.norm.0.forward.output.2', 'torch.float16',
          'torch.float32', [256, 256, 14, 14], [256, 256, 14, 14], True, True,
          0.8, 0.5, 1.1e+10, 1, 0.85, 1, 9, 0.12, 0, 1, 1, 0.1, 1, 1,
          True, 'Yes', '', 'output', 'Functional.batch.norm.0.forward']

base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'test_highlight')


def generate_result_xlsx(base_dir):
    data_path = os.path.join(base_dir, 'target_result.xlsx')
    data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
             'torch.float32', 'torch.float32', [2, 2], [2, 2], 'True', 'True',
             '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, 'True', 'Yes', '', '-1']
            ]
    columns = CompareConst.COMPARE_RESULT_HEADER + ['Data_name']
    result_df = pd.DataFrame(data, columns=columns)
    result_df.to_excel(data_path, index=False, sheet_name='Sheet')
    wb = load_workbook(data_path)
    ws = wb.active
    red_fill = PatternFill(start_color=CompareConst.RED, end_color=CompareConst.RED, fill_type='solid')
    for row_index, row in enumerate(ws.iter_rows()):
        if row_index == 0:
            continue
        for cell in row:
            cell.fill = red_fill
    wb.save(data_path)

    data_path_yellow = os.path.join(base_dir, 'target_result_yellow.xlsx')
    result_df.to_excel(data_path_yellow, index=False, sheet_name='Sheet')
    wb = load_workbook(data_path_yellow)
    ws = wb.active
    yellow_fill = PatternFill(start_color=CompareConst.YELLOW, end_color=CompareConst.YELLOW, fill_type='solid')
    for row_index, row in enumerate(ws.iter_rows()):
        if row_index == 0:
            continue
        for cell in row:
            cell.fill = yellow_fill
    wb.save(data_path_yellow)


def compare_excel_files_with_highlight(file1, file2):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    if len(wb1.sheetnames) != len(wb2.sheetnames):
        return False

    for sheet_name in wb1.sheetnames:
        if sheet_name not in wb2.sheetnames:
            return False
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]

        for row_index, row in enumerate(ws1.iter_rows()):
            if row_index == 0:
                continue
            for cell in row:
                other_cell = ws2[cell.coordinate]

                if cell.value != other_cell.value:
                    return False
                if cell.fill.start_color.index != other_cell.fill.start_color.index:
                    return False
    return True


class TestUtilsMethods(unittest.TestCase):

    def setUp(self):
        os.makedirs(base_dir, mode=0o750, exist_ok=True)

    def tearDown(self):
        if os.path.exists(base_dir):
            shutil.rmtree(base_dir)

    def test_CheckOrderMagnitude_normal(self):
        api_in = [1, 1, 1, 1, 1, 1, True, True, 5, 1, 1]
        api_out = [1, 1, 1, 1, 1, 1, True, True, 1, 1, 1]
        info = (api_in, api_out, 1)
        color_columns = ()
        dump_mode = Const.SUMMARY

        result = CheckOrderMagnitude().apply(info, color_columns, dump_mode)

        self.assertEqual(result, None)

    def test_CheckOneThousandErrorRatio_str(self):
        api_in = [1, 1, 1, 1, 1, 1, True, True, 0.9, 0.5, 1, 1, "unsupported"]
        api_out = [1, 1, 1, 1, 1, 1, True, True, 0.9, 0.5, 1, 1, "unsupported"]
        info = (api_in, api_out, 1)
        color_columns = ()
        dump_mode = Const.ALL

        result = CheckOneThousandErrorRatio().apply(info, color_columns, dump_mode)

        self.assertEqual(result, None)

    @patch("msprobe.core.compare.highlight.add_highlight_row_info")
    def test_CheckOneThousandErrorRatio_red(self, mock_add_highlight_row_info):
        api_in = [1, 1, 1, 1, 1, 1, True, True, 0.9, 0.5, 1, 1, 1]
        api_out = [1, 1, 1, 1, 1, 1, True, True, 0.9, 0.5, 1, 1, 0.5]
        info = (api_in, api_out, 1)
        ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])
        color_columns = ColorColumns(red=[], yellow=[])
        dump_mode = Const.ALL

        CheckOneThousandErrorRatio().apply(info, color_columns, dump_mode)

        mock_add_highlight_row_info.assert_called_once()

    def test_CheckCosineSimilarity_str(self):
        api_in = [1, 1, 1, 1, 1, 1, True, True, "unsupported", 1, 1, "unsupported"]
        api_out = [1, 1, 1, 1, 1, 1, True, True, "unsupported", 1, 1, "unsupported"]
        info = (api_in, api_out, 1)
        color_columns = ()
        dump_mode = Const.ALL

        result = CheckCosineSimilarity().apply(info, color_columns, dump_mode)

        self.assertEqual(result, None)

    def test_CheckMaxRelativeDiff_red(self):
        ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])

        red_lines, yellow_lines = [], []
        color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)

        api_in = {8: 0, 20: 1}
        api_out = {8: 0.6, 20: 1}
        num = 1
        info = (api_in, api_out, num)
        CheckMaxRelativeDiff().apply(info, color_columns, dump_mode=Const.SUMMARY)
        red_lines, yellow_lines = [(1, ["maximum relative error exceeds 0.5"])], []
        target_color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)
        self.assertEqual(color_columns, target_color_columns)

    def test_CheckMaxRelativeDiff_yellow(self):
        ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])

        red_lines, yellow_lines = [], []
        color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)

        api_in = {8: 0.001, 20: 1}
        api_out = {8: 0.2, 20: 1}
        num = 1
        info = (api_in, api_out, num)
        CheckMaxRelativeDiff().apply(info, color_columns, dump_mode=Const.SUMMARY)
        red_lines, yellow_lines = [], [(1, ["The output's maximum relative error exceeds 0.1, while the input/parameter's is below 0.01"])]
        target_color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)
        self.assertEqual(color_columns, target_color_columns)

    def test_CheckMaxRelativeDiff_other_type(self):
        ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])

        red_lines, yellow_lines = [], []
        color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)

        api_in = {8: 0.001, 20: np.nan}
        api_out = {8: 0.2, 20: 1}
        num = 1
        info = (api_in, api_out, num)
        result = CheckMaxRelativeDiff().apply(info, color_columns, dump_mode=Const.SUMMARY)
        self.assertEqual(result, None)

    def test_find_error_rows_normal(self):
        compare_result = np.array([
            ["Functional.linear.0.forward.input.0", "Functional.linear.0.forward.input.0",
             "torch.float32", "torch.float32", [2, 2], [2, 2], 'True', 'True',
             0.0, 0.0, 0.0, 0.0, "0.0%", "0.0%", "0.0%", "0.0%",
             1, 1, 1, 1, 1, 1, 1, 1, True, "", ""],
            ["Functional.linear.0.forward.input.1", "Functional.linear.0.forward.input.1",
             "torch.float32", "torch.float32", [2, 2], [2, 2], 'True', 'True',
             0.0, 0.0, 0.0, 0.0, "0.0%", "0.0%", "0.0%", "0.0%",
             1, 1, 1, 1, 1, 1, 1, 1, True, "", ""],
            ["Functional.linear.0.forward.input.2", "Functional.linear.0.forward.input.2",
             "torch.float32", "torch.float32", [2], [2], 'True', 'True',
             0.0, 0.0, 0.0, 0.0, "0.0%", "0.0%", "0.0%", "0.0%",
             1, 1, 1, 1, 1, 1, 1, 1, True, "", ""],
            ["Functional.linear.0.forward.output.0", "Functional.linear.0.forward.output.0",
             "torch.float32", "torch.float32", [2, 2], [2, 2], 'True', 'True',
             0.0, 0.0, 0.0, 0.0, "0.0%", "0.0%", "0.0%", "0.0%",
             1, 1, 1, 1, 1, 1, 1, 1, True, "", ""],
        ], dtype=object)
        api_batch = ApiBatch("Functional.linear.0.forward", 0)
        api_batch.input_len = 3
        api_batch.output_end_index = 4
        api_batch.params_end_index = 4
        highlight_dict = {"red_lines": [], "red_rows": set(), "yellow_lines": [], "yellow_rows": set()}

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.find_error_rows(compare_result, api_batch, highlight_dict)

        self.assertEqual(highlight_dict, {"red_lines": [], "red_rows": set(), "yellow_lines": [], "yellow_rows": set()})

    def test_find_error_rows_md5(self):
        compare_result = []
        api_batch = ApiBatch("", 0)
        api_batch.input_len = 0
        api_batch.output_end_index = 1
        api_batch.params_end_index = 1
        highlight_dict = {}

        mode_config = ModeConfig(dump_mode=Const.MD5)
        highlight = HighLight(mode_config)
        result = highlight.find_error_rows(compare_result, api_batch, highlight_dict)

        self.assertEqual(result, None)

    @patch("msprobe.core.compare.highlight.logger")
    def test_value_check(self, mock_logger):
        value = "=functional.conv2d"
        api_name = "=functional.conv2d"
        i = 1
        result_df_columns = CompareConst.COMPARE_RESULT_HEADER

        mode_config = ModeConfig()
        highlight = HighLight(mode_config)
        highlight.value_check(value, api_name, i, result_df_columns)

        mock_logger.error.assert_called_once_with(
            "Malicious value [=functional.conv2d] at api_name [=functional.conv2d], column [Bench Name], "
            "is not allowed to be written into the compare result xlsx."
        )

    def test_df_malicious_value_check(self):
        columns = CompareConst.COMPARE_RESULT_HEADER
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '']
                ]
        result_df = pd.DataFrame(data, columns=columns)

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.df_malicious_value_check(result_df, columns)

    def test_compare_result_df_convert(self):
        value = float("nan")
        mode_config = ModeConfig()
        highlight = HighLight(mode_config)
        result = highlight.compare_result_df_convert(value)
        self.assertEqual(result, "nan\t")

    def test_highlight_rows_xlsx_red(self):
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1']
                ]
        columns = CompareConst.COMPARE_RESULT_HEADER + ['Data_name']
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {'red_rows': [0]}
        file_path = os.path.join(base_dir, 'result.xlsx')

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.highlight_rows_xlsx(result_df, highlight_dict, file_path)

        generate_result_xlsx(base_dir)
        self.assertTrue(compare_excel_files_with_highlight(file_path, os.path.join(base_dir, 'target_result.xlsx')))

    def test_highlight_rows_xlsx_yellow(self):
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1']
                ]
        columns = CompareConst.COMPARE_RESULT_HEADER + ['Data_name']
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {'yellow_rows': [0]}
        file_path = os.path.join(base_dir, 'result.xlsx')

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.highlight_rows_xlsx(result_df, highlight_dict, file_path)

        generate_result_xlsx(base_dir)
        self.assertTrue(compare_excel_files_with_highlight(file_path, os.path.join(base_dir, 'target_result_yellow.xlsx')))

    @patch("msprobe.core.compare.highlight.save_workbook")
    def test_highlight_rows_xlsx_malicious_columns(self, mock_save_book):
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1']
                ]
        columns = CompareConst.COMPARE_RESULT_HEADER + ['=Data_name']
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {}
        file_path = base_dir

        temp_output_file = 'temp_output.txt'
        sys.stdout = open(temp_output_file, 'w')

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.highlight_rows_xlsx(result_df, highlight_dict, file_path)

        with open(temp_output_file, 'r') as f:
            output = f.read()
        os.remove(temp_output_file)

        self.assertIn('Malicious value [=Data_name]', output)

    @patch("msprobe.core.compare.highlight.save_workbook")
    def test_highlight_rows_xlsx_malicious_type(self, mock_save_book):
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 '=torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1'],
                ['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 '=torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1']
                ]
        columns = CompareConst.COMPARE_RESULT_HEADER + ['Data_name']
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {'red_rows': [], 'yellow_rows': []}
        file_path = base_dir

        temp_output_file = 'temp_output.txt'
        sys.stdout = open(temp_output_file, 'w')

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.highlight_rows_xlsx(result_df, highlight_dict, file_path)

        with open(temp_output_file, 'r') as f:
            output = f.read()
        os.remove(temp_output_file)

        self.assertIn('Malicious value [=torch.float32]', output)

    def test_add_highlight_row_info_existing(self):
        color_list = [(1, ["a", "b"]), (5, ["c"])]
        num = 5
        highlight_err_msg = "highlight"
        add_highlight_row_info(color_list, num, highlight_err_msg)
        self.assertEqual(color_list, [(1, ["a", "b"]), (5, ["c", "highlight"])])

    def test_update_highlight_err_msg(self):
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1'],
                ['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                 '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', '', '-1']
                ]
        columns = CompareConst.COMPARE_RESULT_HEADER + ['Data_name']
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {
            'red_rows': set([0]),
            'yellow_rows': {0, 1},
            'red_lines': [(0, ['a', 'b'])],
            'yellow_lines': [(0, ['c']), (1, ['d'])]
        }

        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.update_highlight_err_msg(result_df, highlight_dict)

        t_data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                   'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                   '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', 'a\nb', '-1'],
                  ['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                   'torch.float32', 'torch.float32', [2, 2], [2, 2], True, True,
                   '', '', '', '', '', '', 1, 1, 1, 1, 1, 1, 1, 1, True, 'Yes', 'd', '-1']
                  ]
        target_result_df = pd.DataFrame(t_data, columns=columns)
        self.assertTrue(result_df.equals(target_result_df))

    def test_update_highlight_err_msg_md5(self):
        data = [['Functional.linear.0.forward.input.0', 'Functional.linear.0.forward.input.0',
                 'torch.float32', 'torch.float32', True, True, [2, 2], [2, 2], 'abc', 'abc', True, 'pass']
                ]
        columns = CompareConst.MD5_COMPARE_RESULT_HEADER
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {}

        mode_config = ModeConfig(dump_mode=Const.MD5)
        highlight = HighLight(mode_config)
        result = highlight.update_highlight_err_msg(result_df, highlight_dict)

        self.assertEqual(result, None)

    def test_update_highlight_err_msg_fail(self):
        data = [
            ['err_msg1'],
            ['err_msg2']
        ]
        columns = ['Err_message']
        result_df = pd.DataFrame(data, columns=columns)
        highlight_dict = {
            'red_rows': set([0]),
            'yellow_rows': {0, 1},
            'red_lines': [(0, ['a', 'b'])],
            'yellow_lines': [(0, ['c']), (1, ['d'])]
        }
        mode_config = ModeConfig()
        highlight = HighLight(mode_config)
        result = highlight.update_highlight_err_msg(result_df, highlight_dict)
        self.assertEqual(result, None)

    def test_find_error_rows(self):
        api_batch = ApiBatch("Functional_batch_norm_0_forward", 0)
        api_batch.input_len = 1
        api_batch.output_end_index = 4
        api_batch.params_end_index = 4
        summary_result = [summary_line_input, summary_line_1, summary_line_2, summary_line_3]
        highlight_dict_test = {"red_rows": set(), "yellow_rows": set(), "red_lines": [], "yellow_lines": []}
        mode_config = ModeConfig()
        highlight = HighLight(mode_config)
        highlight.find_error_rows(summary_result, api_batch, highlight_dict_test)
        self.assertEqual(highlight_dict_test,
                         {"red_rows": set(), "yellow_rows": set(), "red_lines": [], "yellow_lines": []})

    def test_find_compare_result_error_rows(self):
        result = [line_input, line_1, line_2, line_3]
        result_df = pd.DataFrame(result)
        highlight_dict_test = {"red_rows": set(), "yellow_rows": set(), "red_lines": [], "yellow_lines": []}
        mode_config = ModeConfig(dump_mode=Const.ALL)
        highlight = HighLight(mode_config)
        highlight.find_compare_result_error_rows(result_df, highlight_dict_test)
        self.assertEqual(highlight_dict_test, {
            "red_rows": {1, 3},
            "yellow_rows": {2},
            "red_lines": [
                (1, ["maximum or minimum is nan, -inf, or inf"]),
                (3, ["maximum absolute error exceeds 1e+10"])
            ],
            "yellow_lines": [
                (2, ["The output's one thousandth err ratio decreases by more than 0.1 compared to the input/parameter's"]),
                (3, [
                    "maximum absolute error of both input/parameters and output exceed 1, "
                    "with the output larger by an order of magnitude",
                    "The output's cosine decreases by more than 0.1 compared to the input/parameter's"])
            ]
        })
