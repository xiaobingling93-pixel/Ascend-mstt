#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
# Copyright (C) 2019-2020. Huawei Technologies Co., Ltd. All rights reserved.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""

import json
import math
import multiprocessing
import os.path
import stat
import sys
import warnings
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from collections import namedtuple

from .match import graph_mapping
from ..advisor.advisor import Advisor
from ..common.utils import check_compare_param, add_time_with_xlsx, \
    print_info_log, print_warn_log, print_error_log, CompareException, Const, \
    CompareConst, format_value, check_file_not_exists, check_configuration_param, \
    is_summary_compare, is_md5_compare
from ..common.file_check_util import FileChecker, FileCheckConst, change_mode, FileOpen


def correct_data(result):
    if result == CompareConst.NAN:
        return result
    if float(result) > 0.99999:
        return 1.0
    return result


def cosine_similarity(n_value, b_value):
    np.seterr(divide='ignore', invalid='ignore')
    if len(n_value) == 1:
        return "unsupported", "This tensor is scalar."
    num = n_value.dot(b_value)
    a_norm = np.linalg.norm(n_value)
    b_norm = np.linalg.norm(b_value)
    message = ''
    if a_norm <= Const.FLOAT_EPSILON and b_norm <= Const.FLOAT_EPSILON:
        result = 1.0
    elif a_norm <= Const.FLOAT_EPSILON:
        message = 'Cannot compare by Cosine Similarity, All the data is Zero in npu dump data.'
        result = CompareConst.NAN
    elif b_norm <= Const.FLOAT_EPSILON:
        message = 'Cannot compare by Cosine Similarity, All the data is Zero in Bench dump data.'
        result = CompareConst.NAN
    else:
        cos = num / (a_norm * b_norm)
        if np.isnan(cos):
            message = 'Cannot compare by Cosine Similarity, the dump data has NaN.'
            result = CompareConst.NAN
        else:
            result = format_value(cos)
    result = correct_data(result)
    return result, message


def get_rmse(n_value, b_value):
    if len(n_value) == 0 and len(b_value) == 0:
        rmse = '0'
    elif len(n_value) == 0:
        rmse = CompareConst.NAN
    elif len(b_value) == 0:
        rmse = CompareConst.NAN
    else:
        rmse = np.linalg.norm(n_value - b_value) / np.sqrt(len(n_value))
    if np.isnan(rmse):
        rmse = CompareConst.NAN
    return rmse, ""


def get_mape(n_value, b_value):
    if len(n_value) == 0 and len(b_value) == 0:
        mape = '0'
    elif len(n_value) == 0:
        mape = CompareConst.NAN
    elif len(b_value) == 0:
        mape = CompareConst.NAN
    elif not np.all(n_value) and not np.all(b_value):
        mape = '0'
    elif not np.all(b_value):
        mape = CompareConst.NAN
    else:
        mape_val = np.sum(np.abs((n_value - b_value) / b_value)) / len(b_value) * 100
        mape = CompareConst.NAN if np.isnan(mape_val) else str(round(mape_val, 4)) + '%'
    return mape, ""


def get_max_abs_err(n_value, b_value):
    temp_res = n_value - b_value
    max_value = np.max(np.abs(temp_res))
    return format_value(max_value), ""


def get_relative_err(n_value, b_value):
    np.seterr(divide='ignore', invalid='ignore')
    if b_value.dtype in CompareConst.FLOAT_TYPE:
        zero_mask = (b_value == 0)
        b_value[zero_mask] += np.finfo(b_value.dtype).eps
        n_value[zero_mask] += np.finfo(b_value.dtype).eps
    else:
        n_value, b_value = n_value.astype(float), b_value.astype(float)
        zero_mask = (b_value == 0)
        b_value[zero_mask] += np.finfo(float).eps
        n_value[zero_mask] += np.finfo(float).eps
    relative_err = np.divide((n_value - b_value), b_value)
    return np.abs(relative_err)


def get_max_relative_err(n_value, b_value, input_relative_err=None):
    if input_relative_err is None:
        relative_err = get_relative_err(n_value, b_value)
    else:
        relative_err = input_relative_err
    max_relative_err = np.max(np.abs(relative_err))
    if np.isnan(max_relative_err):
        message = 'Cannot compare by MaxRelativeError, the data contains nan in dump data.'
        return CompareConst.NAN, message
    return format_value(max_relative_err), ""


def rel_err_ratio(relative_err, threshold):
    return format_value(np.sum(relative_err < threshold) / np.size(relative_err))


def check_graph_mode(a_op_name, b_op_name):
    if "Aten" in a_op_name and "Aten" not in b_op_name:
        return True
    if "Aten" not in a_op_name and "Aten" in b_op_name:
        return True
    return False


def check_op(npu_dict, bench_dict, fuzzy_match):
    a_op_name = npu_dict["op_name"]
    b_op_name = bench_dict["op_name"]
    graph_mode = check_graph_mode(a_op_name[0], b_op_name[0])
    if graph_mode:
        return graph_mapping.match(a_op_name[0], b_op_name[0])
    struct_match = check_struct_match(npu_dict, bench_dict)
    if not fuzzy_match:
        return a_op_name == b_op_name and struct_match
    is_match = True
    try:
        is_match = fuzzy_check_op(a_op_name, b_op_name)
    except Exception as err:
        print_warn_log("%s and %s can not fuzzy match." % (a_op_name, b_op_name))
        is_match = False
    return is_match and struct_match


def check_struct_match(npu_dict, bench_dict):
    npu_struct_in = npu_dict.get("input_struct")
    bench_struct_in = bench_dict.get("input_struct")
    npu_struct_out = npu_dict.get("output_struct")
    bench_struct_out = bench_dict.get("output_struct")
    is_match = npu_struct_in == bench_struct_in and npu_struct_out == bench_struct_out
    if not is_match:
        if len(npu_struct_in) == 0 or len(bench_struct_in) == 0 or len(npu_struct_in) != len(bench_struct_in):
            return False
        struct_in_is_match = check_type_shape_match(npu_struct_in, bench_struct_in)
        struct_out_is_match = check_type_shape_match(npu_struct_out, bench_struct_out)
        is_match = struct_in_is_match and struct_out_is_match
    return is_match


def check_type_shape_match(npu_struct, bench_struct):
    shape_type_match = False
    for npu_type_shape, bench_type_shape in zip(npu_struct, bench_struct):
        npu_type = npu_type_shape[0]
        npu_shape = npu_type_shape[1]
        bench_type = bench_type_shape[0]
        bench_shape = bench_type_shape[1]
        shape_match = npu_shape == bench_shape
        type_match = npu_type == bench_type
        if not type_match:
            if [npu_type, bench_type] in [["torch.float16", "torch.float32"], ["torch.float32", "torch.float16"],
                                          ["torch.float16", "torch.bfloat16"], ["torch.bfloat16", "torch.float16"]]:
                type_match = True
            else:
                type_match = False
        shape_type_match = shape_match and type_match
        if not shape_type_match:
            return False
    return shape_type_match


def fuzzy_check_op(npu_name_list, bench_name_list):
    if len(npu_name_list) == 0 or len(bench_name_list) == 0 or len(npu_name_list) != len(bench_name_list):
        return False
    is_match = True
    for npu_name, bench_name in zip(npu_name_list, bench_name_list):
        is_match = fuzzy_check_name(npu_name, bench_name)
        if not is_match:
            break
    return is_match


def fuzzy_check_name(npu_name, bench_name):
    if "forward" in npu_name and "forward" in bench_name:
        is_match = rename_api(npu_name, "forward") == rename_api(bench_name, "forward")
    elif "backward" in npu_name and "backward" in bench_name:
        is_match = rename_api(npu_name, "backward") == rename_api(bench_name, "backward")
    else:
        is_match = npu_name == bench_name
    return is_match


def rename_api(npu_name, process):
    npu_split = npu_name.split(process)
    torch_func_index, in_out = npu_split[0], npu_split[1]
    torch_func_split = torch_func_index.rsplit("_", 2)
    torch_func = str(torch_func_split[0]) + str(in_out)
    return torch_func


def merge_tensor(tensor_list):
    """Return a dictionary of one API with its name, input, output, summery and stack info"""
    op_dict = {}
    op_dict["op_name"] = []
    op_dict["input_struct"] = []
    op_dict["output_struct"] = []
    op_dict["summery"] = []
    op_dict["stack_info"] = []

    for tensor in tensor_list:
        if tensor[0].find("stack_info") != -1:
            if len(tensor) != Const.STACK_COLUMN_NUM:
                print_error_log(f"This stack_info data is not complete. {tensor}")
                raise CompareException(CompareException.INVALID_DATA_ERROR)
            op_dict["stack_info"].append(tensor[1])
            break
        op_dict["op_name"].append(tensor[0])
        if len(tensor) != Const.SUMMARY_COLUMN_NUM:
            print_error_log(f"This summary data is not complete. {tensor}")
            raise CompareException(CompareException.INVALID_DATA_ERROR)
        if tensor[0].find("input") != -1:
            op_dict["input_struct"].append((tensor[3], tensor[4], tensor[2]))
        elif tensor[0].find("output") != -1:
            op_dict["output_struct"].append((tensor[3], tensor[4], tensor[2]))

        if tensor[1] <= Const.DUMP_RATIO_MAX:
            op_dict["summery"].append(tensor[5])

    return op_dict if op_dict["op_name"] else {}


def read_op(ops_queue, pkl_file_handle, stack_mode):
    """Read input pkl file and get one API info with dictionary format once"""
    tensor_list = []
    read_err = False
    read_output_flag = {"last_line": False, "curr_line": False}
    end_flag = "stack_info" if stack_mode is True else "output"

    while True:
        curr_pos = pkl_file_handle.tell()
        tensor_line = pkl_file_handle.readline()
        if len(tensor_line) == 0 and not read_output_flag.get("curr_line"):
            read_err = True
            break
        if tensor_line == '\n':
            continue
        if len(tensor_line) != 0:
            tensor_data = json.loads(tensor_line)
            if not isinstance(tensor_data, list):
                print_error_log(f"This data is not a list, please check the dump data pkl file. {tensor_data}")
                raise CompareException(CompareException.INVALID_DATA_ERROR)
            read_output_flag["last_line"] = read_output_flag.get("curr_line")
            read_output_flag["curr_line"] = True if tensor_data[0].find(end_flag) != -1 else False

        if (read_output_flag.get("last_line") and not read_output_flag.get("curr_line")) \
                or (len(tensor_line) == 0 and read_output_flag.get("curr_line")):  # end of file scenario
            merge_list = merge_tensor(tensor_list)
            if merge_list:
                ops_queue.append(merge_list)
            # the pos of the handle needs to restore to the start of the next api.
            pkl_file_handle.seek(curr_pos, 0)
            break
        tensor_list.append(tensor_data)

    return not read_err


def match_op(npu_queue, bench_queue, fuzzy_match):
    """Match NPU with bench API"""
    for b_index, b_op in enumerate(bench_queue[0: -1]):
        if check_op(npu_queue[-1], b_op, fuzzy_match):
            return len(npu_queue) - 1, b_index
    if check_op(npu_queue[-1], bench_queue[-1], fuzzy_match):
        return len(npu_queue) - 1, len(bench_queue) - 1
    for n_index, n_op in enumerate(npu_queue[0: -1]):
        if check_op(n_op, bench_queue[-1], fuzzy_match):
            return n_index, len(bench_queue) - 1
    return -1, -1


def get_accuracy(result, n_dict, b_dict, highlight_dict, summary_compare=False, md5_compare=False):
    """Compare and add one API's input and output results to dataframe"""
    def get_accuracy_core(n_start, n_len, b_start, b_len, key):
        min_len = min(n_len, b_len)
        npu_stack_info = n_dict.get("stack_info", None)
        bench_stack_info = b_dict.get("stack_info", None)
        has_stack = npu_stack_info and bench_stack_info
        for index in range(min_len):
            n_name = n_dict['op_name'][n_start + index]
            b_name = b_dict['op_name'][b_start + index]
            n_struct = n_dict[key][index]
            b_struct = b_dict[key][index]
            err_msg = ""
            if md5_compare:
                result_item = [n_name, b_name, n_struct[0], b_struct[0], n_struct[1], b_struct[1],
                               n_struct[2], b_struct[2], CompareConst.PASS if n_struct[2] == b_struct[2] else CompareConst.DIFF]
                if has_stack and index == 0 and key == "input_struct":
                    result_item.extend(npu_stack_info)
                result.append(result_item)
                continue

            if summary_compare:
                result_item = [n_name, b_name, n_struct[0], b_struct[0], n_struct[1], b_struct[1],
                               " ", " ", " ", " "]
            else:
                result_item = [n_name, b_name, n_struct[0], b_struct[0], n_struct[1], b_struct[1],
                               " ", " ", " ", " ", " "]

            npu_summery_data = n_dict.get("summery")[n_start + index]
            result_item.extend(npu_summery_data)
            bench_summery_data = b_dict.get("summery")[b_start + index]
            result_item.extend(bench_summery_data)

            if summary_compare:
                start_idx = CompareConst.SUMMARY_COMPARE_RESULT_HEADER.index(CompareConst.MAX_DIFF)
                warning_flag = False
                for i, (npu_val, bench_val) in enumerate(zip(npu_summery_data, bench_summery_data)):
                    if isinstance(npu_val, (float, int)) and isinstance(bench_val, (float, int)):
                        diff = npu_val - bench_val
                        result_item[start_idx + i] = diff
                        magnitude_diff = abs(diff) / (max(abs(npu_val), abs(bench_val)) + 1e-10)
                        if magnitude_diff > 0.5:
                            warning_flag = True
                    else:
                        result_item[start_idx + i] = CompareConst.NAN
                accuracy_check = CompareConst.WARNING if warning_flag else ""
                err_msg += "Need double check api accuracy." if warning_flag else ""
                result_item[start_idx:] = [f'{str(x)}\t' if str(x) in ('inf', '-inf', 'nan') else x for x in result_item[start_idx:]]

            result_item.append(accuracy_check if summary_compare else CompareConst.ACCURACY_CHECK_YES)
            result_item.append(err_msg)
            if has_stack and index == 0 and key == "input_struct":
                result_item.extend(npu_stack_info)

            result.append(result_item)

        if n_len > b_len:
            for index in range(b_len, n_len):
                n_name = n_dict['op_name'][n_start + index]
                n_struct = n_dict[key][index]
                if md5_compare:
                    result_item = [n_name, CompareConst.NAN, n_struct[0], CompareConst.NAN,
                                   n_struct[1], CompareConst.NAN, n_struct[2], CompareConst.NAN, CompareConst.NAN]
                    result.append(result_item)
                    continue
                result_item = [n_name, CompareConst.NAN, n_struct[0], CompareConst.NAN,
                               n_struct[1], CompareConst.NAN, " ", " ", " ", " ", " "]
                summery_data = n_dict.get("summery")[n_start + index]
                result_item.extend(summery_data)
                summery_data = [CompareConst.NAN for _ in range(len(n_dict.get("summery")[0]))]
                result_item.extend(summery_data)

                err_msg = ""
                result_item.append(CompareConst.ACCURACY_CHECK_YES)
                result_item.append(err_msg)

                if has_stack and index == 0 and key == "input_struct":
                    result_item.extend(npu_stack_info)

                result.append(result_item)

    n_num = len(n_dict['op_name'])
    b_num = len(b_dict['op_name'])
    n_num_input = len([name for name in n_dict['op_name'] if 'input' in name])
    b_num_input = len([name for name in b_dict['op_name'] if 'input' in name])
    n_num_output = n_num - n_num_input
    b_num_output = b_num - b_num_input
    last_len = len(result)
    get_accuracy_core(0, n_num_input, 0, b_num_input, 'input_struct')
    get_accuracy_core(n_num_input, n_num_output, b_num_input, b_num_output, 'output_struct')
    # check whether to highlight API rows if using summary compare
    if summary_compare:
        find_error_rows(result[last_len:], last_len, n_num_input, highlight_dict, summary_compare)


def check_order_magnitude(info, color_columns, summary_compare=True):
    """Check if order magnitude diff of max_diff larger than 1"""
    api_in, api_out, num = info
    max_diff_index = get_header_index('Max diff' if summary_compare else 'MaxAbsErr', summary_compare)
    if abs(api_in[max_diff_index]) > abs(api_out[max_diff_index]):
        return
    in_order = 0 if abs(api_in[max_diff_index]) < 1 else math.log10(abs(api_in[max_diff_index]))
    out_order = 0 if abs(api_out[max_diff_index]) < 1 else math.log10(abs(api_out[max_diff_index]))
    if out_order - in_order >= CompareConst.ORDER_MAGNITUDE_DIFF_YELLOW:
        color_columns.yellow.append(num)


def check_one_thousand_error_ratio(info, color_columns, summary_compare=True):
    """Compare output's one thousand error ratio with input's """
    api_in, api_out, num = info
    one_thousand_index = get_header_index('One Thousandth Err Ratio', summary_compare)
    if not isinstance(api_in[one_thousand_index], (float, int)) or not isinstance(api_out[one_thousand_index], (float, int)):
        return
    if api_in[one_thousand_index] > CompareConst.ONE_THOUSAND_ERROR_IN_RED and api_out[one_thousand_index] < CompareConst.ONE_THOUSAND_ERROR_OUT_RED:
        color_columns.red.append(num)
    elif api_in[one_thousand_index] - api_out[one_thousand_index] > CompareConst.ONE_THOUSAND_ERROR_DIFF_YELLOW:
        color_columns.yellow.append(num)


def check_cosine_similarity(info, color_columns, summary_compare=True):
    """Check if output's cosine similarity more than 0.1 smaller than input's"""
    api_in, api_out, num = info
    cosine_index = get_header_index('Cosine', summary_compare)
    if not isinstance(api_in[cosine_index], (float, int)) or not isinstance(api_out[cosine_index], (float, int)):
        return
    if api_in[cosine_index] - api_out[cosine_index] > CompareConst.COSINE_DIFF_YELLOW:
        color_columns.yellow.append(num)


def check_max_relative_diff(info, color_columns, summary_compare=True):
    """Compare the output value of max_diff / bench_max with input"""
    api_in, api_out, num = info
    max_diff_index = get_header_index('Max diff', summary_compare)
    bench_max_index = get_header_index('Bench max', summary_compare)
    input_max_relative_diff = np.abs(np.divide(api_in[max_diff_index], max(0.01, api_in[bench_max_index])))
    output_max_relative_diff = np.abs(np.divide(api_out[max_diff_index], max(0.01, api_out[bench_max_index])))
    if not isinstance(input_max_relative_diff, (float, int)) or not isinstance(output_max_relative_diff, (float, int)):
        return
    if output_max_relative_diff > CompareConst.MAX_RELATIVE_OUT_RED:
        color_columns.red.append(num)
    elif output_max_relative_diff > CompareConst.MAX_RELATIVE_OUT_YELLOW and input_max_relative_diff < CompareConst.MAX_RELATIVE_IN_YELLOW:
        color_columns.yellow.append(num)


def check_overflow(info, color_columns, summary_compare=False):
    """Check if Inf or Nan exists in NPU max/min, or large number in Max diff"""
    line, num = info
    npu_max_index = get_header_index('NPU max', summary_compare)
    npu_min_index = get_header_index('NPU min', summary_compare)
    max_diff_index = get_header_index('Max diff' if summary_compare else 'MaxAbsErr', summary_compare)
    if str(line[npu_max_index]) in CompareConst.OVERFLOW_LIST or str(line[npu_min_index]) in CompareConst.OVERFLOW_LIST:
        color_columns.red.append(num)
        return
    # check if Max_Diff > 1e+10
    if isinstance(line[max_diff_index], (float, int)) and line[max_diff_index] > CompareConst.MAX_DIFF_RED:
        color_columns.red.append(num)


def get_header_index(header_name, summary_compare=False):
    if summary_compare:
        header = CompareConst.SUMMARY_COMPARE_RESULT_HEADER[:]
    else:
        header = CompareConst.COMPARE_RESULT_HEADER[:]
    if header_name not in header:
        print_error_log(f"{header_name} not in data name")
        raise CompareException(CompareException.INVALID_PARAM_ERROR)
    return header.index(header_name)


class HighlightRules:
    """Highlight rules to check whether API errors"""
    # rules for every line: pass in every line of api to check if error exists
    basic_rules = {
        "check_overflow": check_overflow
    }

    # rules compare output with input: pass in input, output to check if output errors compare to input
    compare_rules = {
        "check_order_magnitude": check_order_magnitude,
        "check_one_thousand_error": check_one_thousand_error_ratio,
        "check_cosine_similarity": check_cosine_similarity
    }
    summary_compare_rules = {
        "check_order_magnitude": check_order_magnitude,
        "check_max_relative_diff": check_max_relative_diff,
    }


def find_error_rows(result, last_len, n_num_input, highlight_dict, summary_compare=False):
    """Find error api and return dict with highlight information red or yellow"""
    npu_max_index = get_header_index('NPU max', summary_compare)
    bench_max_index = get_header_index('Bench max', summary_compare)
    max_diff_index = get_header_index('Max diff' if summary_compare else 'MaxAbsErr', summary_compare)

    red_lines, yellow_lines = [], [] #lines to highlight red or yellow
    LineInfo = namedtuple('LineInfo', ['line_data', 'num_pointer'])
    ApiInfo = namedtuple('ApiInfo', ['api_input', 'api_output', 'num_pointer'])
    ColorColumns = namedtuple('ColorColumns', ['red', 'yellow'])
    color_columns = ColorColumns(red=red_lines, yellow=yellow_lines)

    for i, line in enumerate(result):
        num = last_len + i
        line_info = LineInfo(line_data=line, num_pointer=num)
        for rule in HighlightRules.basic_rules.values():
            rule(line_info, color_columns, summary_compare)

    for n, api_out in enumerate(result[n_num_input:len(result)]):
        num = last_len + n_num_input + n
        if num in red_lines:
            continue
        if not isinstance(api_out[npu_max_index], (float, int)) \
                or not isinstance(api_out[bench_max_index], (float, int)) \
                or not isinstance(api_out[max_diff_index],(float, int)):
            continue
        for m, api_in in enumerate(result[0:n_num_input]):
            if not isinstance(api_in[npu_max_index], (float, int)) \
                    or not isinstance(api_in[bench_max_index], (float, int)) \
                    or not isinstance(api_in[max_diff_index], (float, int)):
                continue

            api_info = ApiInfo(api_input=api_in, api_output=api_out, num_pointer=num)
            if summary_compare:
                for rule in HighlightRules.summary_compare_rules.values():
                    rule(api_info, color_columns, summary_compare)
            else:
                for rule in HighlightRules.compare_rules.values():
                    rule(api_info, color_columns, summary_compare)

    highlight_dict.get('red_rows', []).extend(list(set(red_lines)))
    highlight_dict.get('yellow_rows', []).extend(list(set(yellow_lines) - set(red_lines)))


def get_name_and_state(name):
    """Get api/module name and state"""
    if "input" in name:
        api_name = name.split("input")[0]
        state = "input"
    else:
        api_name = name.split("output")[0]
        state = "output"
    return api_name, state


def find_compare_result_error_rows(result_df, highlight_dict):
    """Group the API with its input and output, then find error API with func find_error_rows"""
    result = result_df.values
    start, input_num, output_num, end = 0, 0, 0, len(result_df)
    last_api_name, last_state = None, None
    num, last_len = 0, 0
    for i in range(len(result)):
        api_name, state = get_name_and_state(result[i][0])
        if last_api_name:
            if api_name == last_api_name:
                if state == last_state:
                    num += 1
                else:
                    input_num = num
                    num, last_state = 1, state
            else:
                output_num = num
                find_error_rows(result[start:start + input_num + output_num], start, input_num, highlight_dict)
                num, last_api_name, last_state = 1, api_name, state
                start += input_num + output_num
                input_num, output_num = 1, 0
        else:
            num, last_api_name, last_state = 1, api_name, state
    if state:
        if state == "input":
            input_num = num
        else:
            output_num = num
        find_error_rows(result[start:start + input_num + output_num], start, input_num, highlight_dict)


def highlight_rows_xlsx(result_df, highlight_dict, file_path):
    """Write and highlight results in Excel"""
    print_info_log('Compare result is %s' % file_path)

    wb = openpyxl.Workbook()
    ws = wb.active

    # write header
    for j, col_name in enumerate(result_df.columns, start=1):
        ws.cell(row=1, column=j, value=col_name)

    for i, row in enumerate(result_df.iterrows(), start=2):
        for j, value in enumerate(row[1], start=1):
            if not isinstance(value, (float, int)):
                value = f'{str(value)}\t' if str(value) in ('inf', '-inf', 'nan') else str(value)
            ws.cell(row=i, column=j, value=f'{str(value)}\t' if str(value) in ('inf', '-inf', 'nan') else value)

            if (i - 2) in highlight_dict['red_rows']:
                ws.cell(row=i, column=j).fill = PatternFill(start_color=CompareConst.RED,
                                                            end_color=CompareConst.RED, fill_type="solid")
            elif (i - 2) in highlight_dict['yellow_rows']:
                ws.cell(row=i, column=j).fill = PatternFill(start_color=CompareConst.YELLOW,
                                                            end_color=CompareConst.YELLOW, fill_type="solid")
    wb.save(file_path)


def _do_multi_process(input_parma, result_df):
    """Use multiprocess to handle real data"""
    try:
        result_df = _handle_multi_process(compare_ops, input_parma, result_df, multiprocessing.Manager().RLock())
        return result_df
    except ValueError as e:
        print_error_log('result dataframe is not found.')
        raise CompareException(CompareException.INVALID_DATA_ERROR) from e


def read_dump_data(result_df):
    """Return a dict to pair npu name with bench name"""
    try:
        npu_dump_name_list = result_df.iloc[0:, 0].tolist()
        bench_dump_name_list = result_df.iloc[0:, 1].tolist()
        op_name_mapping_dict = {}
        for index, _ in enumerate(npu_dump_name_list):
            npu_dump_name = npu_dump_name_list[index]
            bench_dump_name = bench_dump_name_list[index]
            op_name_mapping_dict[npu_dump_name] = [npu_dump_name, bench_dump_name]
        return op_name_mapping_dict
    except ValueError as e:
        print_error_log('result dataframe is not found.')
        raise CompareException(CompareException.INVALID_DATA_ERROR) from e
    except IndexError as e:
        print_error_log('result dataframe elements can not be access.')
        raise CompareException(CompareException.INDEX_OUT_OF_BOUNDS_ERROR) from e


def _handle_multi_process(func, input_parma, result_df, lock):
    """To split and multiprocess real data"""
    process_num = int((multiprocessing.cpu_count() + 1) / 2)
    op_name_mapping_dict = read_dump_data(result_df)

    df_chunk_size = len(result_df) // process_num
    if df_chunk_size > 0:
        df_chunks = [result_df.iloc[i:i + df_chunk_size] for i in range(0, len(result_df), df_chunk_size)]
    else:
        df_chunks = [result_df]

    results =[]
    pool = multiprocessing.Pool(process_num)

    def err_call(args):
        print_error_log('multiprocess compare failed! Reason: {}'.format(args))
        try:
            pool.terminate()
        except OSError as e:
            print_error_log("pool terminate failed")

    for process_idx, df_chunk in enumerate(df_chunks):
        idx = df_chunk_size * process_idx
        result = pool.apply_async(func,
                                  args=(idx, op_name_mapping_dict, df_chunk, lock, input_parma),
                                  error_callback=err_call)
        results.append(result)
    final_results = [r.get() for r in results]
    pool.close()
    pool.join()
    return pd.concat(final_results, ignore_index = True)


def compare_ops(idx, dump_path_dict, result_df, lock, input_parma):
    """Return a dataframe with compare results of real data"""
    cos_result = []
    max_err_result = []
    max_relative_err_result = []
    err_mess = []
    one_thousand_err_ratio_result = []
    five_thousand_err_ratio_result = []
    is_print_compare_log = input_parma.get("is_print_compare_log")
    for i in range(len(result_df)):
        op_name = result_df.iloc[i, 0]
        if is_print_compare_log:
            print("start compare: {}".format(op_name))
        cos_sim, max_abs_err, max_relative_err, err_msg, one_thousand_err_ratio, five_thousand_err_ratio = compare_by_op(op_name, dump_path_dict, input_parma)
        if is_print_compare_log:
            print("[{}] Compare result: cosine {}, max_abs_err {}, max_relative_err {}, {}, one_thousand_err_ratio {}, five_thousand_err_ratio {}".format(op_name, cos_sim, max_abs_err, max_relative_err, err_msg, one_thousand_err_ratio, five_thousand_err_ratio))
        cos_result.append(cos_sim)
        max_err_result.append(max_abs_err)
        max_relative_err_result.append(max_relative_err)
        err_mess.append(err_msg)
        one_thousand_err_ratio_result.append(one_thousand_err_ratio)
        five_thousand_err_ratio_result.append(five_thousand_err_ratio)
    result_df = _save_cmp_result(idx, cos_result, max_err_result, max_relative_err_result, err_mess, one_thousand_err_ratio_result,
                                 five_thousand_err_ratio_result, result_df, lock)
    return result_df

def _save_cmp_result(idx, cos_result, max_err_result, max_relative_err_result, err_msg, one_thousand_err_ratio_result, five_thousand_err_ratio_result, result_df, lock):
    lock.acquire()
    try:
        for i, _ in enumerate(cos_result):
            process_index = i + idx
            result_df.loc[process_index, CompareConst.COSINE] = cos_result[i]
            result_df.loc[process_index, CompareConst.MAX_ABS_ERR] = max_err_result[i]
            result_df.loc[process_index, CompareConst.MAX_RELATIVE_ERR] = max_relative_err_result[i]
            result_df.loc[process_index, CompareConst.ERROR_MESSAGE] = err_msg[i]
            result_df.loc[process_index, CompareConst.ACCURACY] = check_accuracy(cos_result[i], max_err_result[i])
            result_df.loc[process_index, CompareConst.ONE_THOUSANDTH_ERR_RATIO] = one_thousand_err_ratio_result[i]
            result_df.loc[process_index, CompareConst.FIVE_THOUSANDTHS_ERR_RATIO] = five_thousand_err_ratio_result[i]
        return result_df
    except ValueError as e:
        print_error_log('result dataframe is not found.')
        raise CompareException(CompareException.INVALID_DATA_ERROR) from e
    except IndexError as e:
        print_error_log('result dataframe elements can not be access.')
        raise CompareException(CompareException.INDEX_OUT_OF_BOUNDS_ERROR) from e
    finally:
        lock.release()


def check_accuracy(cos, max_abs_err):
    if cos == CompareConst.SHAPE_UNMATCH:
        return CompareConst.ACCURACY_CHECK_UNMATCH
    if cos == CompareConst.NAN or max_abs_err == CompareConst.NAN:
        return CompareConst.NAN
    if cos == "N/A" or max_abs_err == "N/A":
        return CompareConst.ACCURACY_CHECK_NO
    try:
        cos, max_abs_err = float(cos), float(max_abs_err)
    except ValueError:
        print_warn_log("Cosine or MaxAbsErr can not get float value.")
        return CompareConst.NAN
    if cos < CompareConst.COS_THRESHOLD and max_abs_err > CompareConst.MAX_ABS_ERR_THRESHOLD:
        return CompareConst.ACCURACY_CHECK_NO
    if cos < CompareConst.COS_MAX_THRESHOLD or max_abs_err > CompareConst.MAX_ABS_ERR_MAX_THRESHOLD:
        return CompareConst.ACCURACY_CHECK_NO
    return CompareConst.ACCURACY_CHECK_YES


def compare_by_op(op_name, op_name_mapping_dict, input_parma):
    """Compare single NPU data with bench data"""
    npu_bench_name_list = op_name_mapping_dict[op_name]
    if npu_bench_name_list[1] == CompareConst.NAN:
        return CompareConst.NAN, CompareConst.NAN, CompareConst.NAN, CompareConst.NO_BENCH, CompareConst.NAN, CompareConst.NAN
    try:
        n_path = os.path.join(input_parma.get("npu_dump_data_dir"), npu_bench_name_list[0] + ".npy")
        b_path = os.path.join(input_parma.get("bench_dump_data_dir"), npu_bench_name_list[1] + ".npy")
        n_path_checker = FileChecker(n_path, FileCheckConst.FILE, FileCheckConst.READ_ABLE,
                                     FileCheckConst.NUMPY_SUFFIX, False)
        b_path_checker = FileChecker(b_path, FileCheckConst.FILE, FileCheckConst.READ_ABLE,
                                     FileCheckConst.NUMPY_SUFFIX, False)
        n_path = n_path_checker.common_check()
        b_path = b_path_checker.common_check()
        n_value = np.load(n_path)
        b_value = np.load(b_path)
    except IOError as error:
        return CompareConst.NAN, CompareConst.NAN, CompareConst.NAN, "Dump file: {} not found.".format(error.filename), CompareConst.NAN, CompareConst.NAN
    relative_err = get_relative_err(n_value, b_value)
    if len(n_value.shape) == 0:
        if n_value.dtype == bool:
            n_value = n_value.astype(float)
            b_value = b_value.astype(float)
        max_abs_err, _ = get_max_abs_err(n_value, b_value)
        max_relative_err, _ = get_max_relative_err(n_value, b_value, input_relative_err=relative_err)
        return "unsupported", max_abs_err, max_relative_err, "This is type of scalar data, can not compare.", CompareConst.NAN, CompareConst.NAN
    if n_value.size == 0:
        return "unsupported", 0, 0, "This is empty data, can not compare.", 0, 0
    if n_value.shape != b_value.shape:
        return CompareConst.SHAPE_UNMATCH, CompareConst.SHAPE_UNMATCH, CompareConst.SHAPE_UNMATCH, "Shape of NPU and bench Tensor do not match. Skipped.", CompareConst.SHAPE_UNMATCH, CompareConst.SHAPE_UNMATCH
    if n_value.dtype != b_value.dtype:
        print_warn_log("Dtype of NPU and bench Tensor do not match: {}".format(op_name))
        err_msg = " Dtype of NPU and bench Tensor do not match."
    else:
        err_msg = ""

    n_value, b_value = handle_inf_nan(n_value, b_value)
    if n_value is CompareConst.NAN or b_value is CompareConst.NAN:
        return "N/A", "N/A", "N/A", "The position of inf or nan in NPU and bench Tensor do not match.", "N/A", "N/A"

    n_value = n_value.reshape(-1).astype(float)
    b_value = b_value.reshape(-1).astype(float)
    err_msg = ""
    cos_sim, message = cosine_similarity(n_value, b_value)

    abs_err = np.abs(n_value - b_value)
    max_abs_err = format_value(np.max(abs_err))
    max_relative_err, message = get_max_relative_err(n_value, b_value, input_relative_err=relative_err)
    one_thousand_err_ratio = rel_err_ratio(relative_err, 0.001)
    five_thousand_err_ratio = rel_err_ratio(relative_err, 0.005)

    if not err_msg:
        err_msg += message
    else:
        err_msg = err_msg + ' ' + message

    if npu_bench_name_list[0] != npu_bench_name_list[1]:
        err_msg += " Fuzzy matching data, the comparison accuracy may be affected."
    return cos_sim, max_abs_err, max_relative_err, err_msg, one_thousand_err_ratio, five_thousand_err_ratio


def handle_inf_nan(n_value, b_value):
    n_inf = np.isinf(n_value)
    b_inf = np.isinf(b_value)
    n_nan = np.isnan(n_value)
    b_nan = np.isnan(b_value)
    if np.any(n_inf) or np.any(b_inf) or np.any(n_nan) or np.any(b_nan):
        if np.array_equal(n_inf, b_inf) and np.array_equal(n_nan, b_nan):
            n_value[n_inf] = 0
            b_value[b_inf] = 0
            n_value[n_nan] = 0
            b_value[b_nan] = 0
        else:
            return CompareConst.NAN, CompareConst.NAN
    return n_value, b_value


def compare(input_parma, output_path, stack_mode=False, auto_analyze=True,
            fuzzy_match=False):
    try:
        message = Const.VERSION_MESSAGE
        warnings.warn(message)        
        summary_compare = is_summary_compare(input_parma)
        md5_compare = is_md5_compare(input_parma)
        check_configuration_param(stack_mode, auto_analyze, fuzzy_match)
        check_compare_param(input_parma, output_path, stack_mode, summary_compare)
    except CompareException as error:
        print_error_log('Compare failed. Please check the arguments and do it again!')
        sys.exit(error.code)
    compare_core(input_parma, output_path, stack_mode=stack_mode,
                 auto_analyze=auto_analyze, fuzzy_match=fuzzy_match, summary_compare=summary_compare,
                 md5_compare=md5_compare)


def compare_core(input_parma, output_path, stack_mode=False, auto_analyze=True,
                 suffix='', fuzzy_match=False, summary_compare=False, md5_compare=False):
    """Compare NPU with bench data then return compare results and advise"""
    print_info_log("Please check whether the input data belongs to you. If not, there may be security risks.")
    file_name = add_time_with_xlsx("compare_result" + suffix)
    file_path = os.path.join(os.path.realpath(output_path), file_name)
    check_file_not_exists(file_path)
    highlight_dict = {'red_rows': [], 'yellow_rows': []}

    with FileOpen(input_parma.get("npu_pkl_path"), "r") as npu_pkl, \
            FileOpen(input_parma.get("bench_pkl_path"), "r") as bench_pkl:
        result_df = compare_process([npu_pkl, bench_pkl], stack_mode, fuzzy_match, highlight_dict, summary_compare,
                                    md5_compare)

    if not md5_compare and not summary_compare:
        result_df = _do_multi_process(input_parma, result_df)
        find_compare_result_error_rows(result_df, highlight_dict)
    highlight_rows_xlsx(result_df, highlight_dict, file_path)
    change_mode(file_path, FileCheckConst.DATA_FILE_AUTHORITY)
    if auto_analyze:
        advisor = Advisor(result_df, output_path)
        advisor.analysis()


def parse(pkl_file, module_name_prefix):
    if not isinstance(module_name_prefix, str):
        print_error_log("The parameter:module_name_prefix is not a string.")
        raise CompareException(CompareException.INVALID_PARAM_ERROR)
    with FileOpen(pkl_file, "r") as f:
        done = False
        title_printed = False
        while not done:
            pkl_line = f.readline()
            if pkl_line == '\n':
                continue
            if len(pkl_line) == 0:
                done = True
                break

            msg = json.loads(pkl_line)
            info_prefix = msg[0]
            if not info_prefix.startswith(module_name_prefix):
                continue

            if info_prefix.find("stack_info") != -1:
                print("\nTrace back({}):".format(msg[0]))
                for item in reversed(msg[1]):
                    print("  File \"{}\", line {}, in {}".format(item[0], item[1], item[2]))
                    print("    {}".format(item[3]))
                continue
            if len(msg) > 5:
                summery_info = "  [{}][dtype: {}][shape: {}][max: {}][min: {}][mean: {}]" \
                    .format(msg[0], msg[3], msg[4], msg[5][0], msg[5][1], msg[5][2])
                if not title_printed:
                    print("\nStatistic Info:")
                    title_printed = True
                print(summery_info)


def compare_process(file_handles, stack_mode, fuzzy_match, highlight_dict, summary_compare=False, md5_compare=False):
    """Read input pkl files, compare summary/md5 data between NPU and bench and return dataframe"""
    npu_pkl_handle, bench_pkl_handle = file_handles
    if fuzzy_match:
        print_warn_log("This task uses fuzzy matching, which may affect the accuracy of the comparison.")
    npu_ops_queue = []
    bench_ops_queue = []
    result = []
    while True:
        last_npu_ops_len = len(npu_ops_queue)
        npu_file_flag = read_op(npu_ops_queue, npu_pkl_handle, stack_mode)
        bench_file_flag = read_op(bench_ops_queue, bench_pkl_handle, stack_mode)
        if not npu_file_flag and not bench_file_flag:
            break
        if len(npu_ops_queue) == 0 or len(bench_ops_queue) == 0 or len(npu_ops_queue) == last_npu_ops_len:
            continue
        n_match_point, b_match_point = match_op(npu_ops_queue, bench_ops_queue, fuzzy_match)
        if n_match_point == -1 and b_match_point == -1:
            continue
        n_match_data = npu_ops_queue[n_match_point]
        b_match_data = bench_ops_queue[b_match_point]
        un_match_data = npu_ops_queue[0: n_match_point]
        for npu_data in un_match_data:
            get_un_match_accuracy(result, npu_data, md5_compare, summary_compare)
        get_accuracy(result, n_match_data, b_match_data, highlight_dict, summary_compare, md5_compare)
        del npu_ops_queue[0: n_match_point + 1]
        del bench_ops_queue[0: b_match_point + 1]
    if npu_ops_queue:
        for npu_data in npu_ops_queue:
            get_un_match_accuracy(result, npu_data, md5_compare, summary_compare)

    header = []
    if md5_compare:
        header = CompareConst.MD5_COMPARE_RESULT_HEADER[:]
    elif summary_compare:
        header = CompareConst.SUMMARY_COMPARE_RESULT_HEADER[:]
    else:
        header = CompareConst.COMPARE_RESULT_HEADER[:]
    if stack_mode:
        header.append(CompareConst.STACK)
    result_df = pd.DataFrame(result, columns=header)
    return result_df


def get_un_match_accuracy(result, n_dict, md5_compare, summary_compare):
    index_out = 0
    npu_stack_info = n_dict.get("stack_info", None)
    bench_name, bench_type, bench_shape = CompareConst.NAN, CompareConst.NAN, CompareConst.NAN
    err_msg = CompareConst.NO_BENCH
    accuracy_check_res = CompareConst.NAN
    for index, n_name in enumerate(n_dict["op_name"]):
        if n_name.find("input") != -1:
            n_struct = n_dict["input_struct"][index]
        else:
            n_struct = n_dict["output_struct"][index_out]
            index_out += 1

        result_item = [n_name, bench_name, n_struct[0], bench_type, n_struct[1], bench_shape]
        if md5_compare:
            result_item.extend([CompareConst.NAN] * 3)
            if npu_stack_info and index == 0:
                result_item.extend(npu_stack_info)
            result.append(result_item)
            continue
        if summary_compare:
            result_item.extend([CompareConst.NAN] * 4)
        else:
            result_item.extend([CompareConst.NAN] * 5)
        summery_data = n_dict.get("summery")[index]
        result_item.extend(summery_data)
        summery_data = [CompareConst.NAN] * 4
        result_item.extend(summery_data)
        result_item.append(accuracy_check_res)
        result_item.append(err_msg)
        if npu_stack_info and index == 0:
            result_item.extend(npu_stack_info)
        result.append(result_item)
